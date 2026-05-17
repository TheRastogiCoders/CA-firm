"""
Regenerate WEBSITE_CONTENT_AUDIT.xlsx from live frontend source (user-visible text).
"""
import json
import re
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
DATA_JSON = Path(__file__).resolve().parent / 'audit_data.json'
OUT_PATHS = [
    ROOT / 'WEBSITE_CONTENT_AUDIT.xlsx',
    Path(r'c:\Users\vasur\Downloads\WEBSITE_CONTENT_AUDIT.xlsx'),
]

FONT_BODY = 'DM Sans, 400, ~1rem (body)'
FONT_EYEBROW = 'DM Sans, 600, ~0.8125rem (section eyebrow)'
FONT_TITLE = 'DM Sans, 600, clamp(1.5–2.25rem) (.about-title / .services-title)'
FONT_HERO_TITLE = 'DM Sans, 600, clamp(1.95–3.25rem) (.hero-title)'
FONT_HERO_TAG = 'DM Sans, 500–600, clamp(0.92–1rem) (.hero-tagline)'
FONT_PAGE_TITLE = 'DM Sans, 600, clamp(1.9–3rem) (.page-title)'
FONT_CARD = 'DM Sans, 600, ~1.0625–1.25rem (cards)'
REMARKS = 'Auto-generated from frontend source. Column H = typography on live site (DM Sans).'


class AuditBuilder:
    def __init__(self):
        self.rows = []
        self.sno = 0
        self._page = None
        self._page_no = None

    def add(
        self,
        page_name,
        page_no,
        subheading,
        design,
        buttons,
        element,
        text,
        font=FONT_BODY,
        remarks=None,
    ):
        self.sno += 1
        show_page = page_name if page_name != self._page else None
        if page_name:
            self._page = page_name
            self._page_no = page_no
        self.rows.append(
            [
                self.sno,
                show_page,
                page_no if show_page else None,
                subheading,
                design,
                buttons or None,
                element,
                font,
                text,
                remarks or REMARKS,
            ]
        )

    def section(self, page, page_no, sub, design='', buttons=''):
        self._sec = (page, page_no, sub, design, buttons)

    def el(self, element, text, font=FONT_BODY, buttons=None):
        p, pn, sub, design, btns = self._sec
        self.add(p, pn, sub, design, buttons or btns, element, text, font, None)


def load_data():
    with open(DATA_JSON, encoding='utf-8') as f:
        return json.load(f)


def build_header(builder):
    builder.section('HEADER (all pages)', 0, 'Info ticker', 'Header ticker', '')
    for item in [
        'Email: shikhar.dwivedi@dgc.ind.in',
        'Mobile: +91 9721227799',
        'Landline: 0542-2502525',
        'WhatsApp: +91 9721227799',
        'Head Office: Varanasi',
        'Branches: Delhi | Kolkata | Bokaro',
        'RBI Registered & CAG Empanelled',
        'Mon-Sat: 10:00 AM - 7:00 PM',
    ]:
        builder.el('Ticker item', item, FONT_BODY)
    builder.section('HEADER (all pages)', 0, 'Brand', 'Header', '')
    builder.el('Firm name', 'DWIVEDI GUPTA & Co.', FONT_CARD)
    builder.el('Tagline', 'Chartered Accountants', FONT_BODY)
    builder.section('HEADER (all pages)', 0, 'Navigation', 'Nav', '')
    for link in [
        'Home', 'About Us', 'Services', 'Industries', 'Team', 'Clients', 'Insights', 'Contact',
    ]:
        builder.el('Nav link', link, FONT_BODY)
    builder.el('CTA button', 'Schedule Consultation', FONT_BODY, 'Schedule Consultation')
    builder.section('HEADER (all pages)', 0, 'Services dropdown', 'Dropdown', '')
    for s in [
        'All Services', 'Tax & Compliance', 'Audit & Assurance', 'GST Advisory', 'Corporate Law',
        'Company Formation', 'Financial Consulting', 'Project Finance', 'Government Schemes Advisory',
    ]:
        builder.el('Menu item', s, FONT_BODY)
    builder.section('HEADER (all pages)', 0, 'Team dropdown', 'Dropdown', '')
    for t in ['Partners', 'Team Members', 'Support Team']:
        builder.el('Menu item', t, FONT_BODY)


def build_home(builder):
    slides = [
        ('Chartered Accountants | Advisors | Consultants', 'Chartered Accountants for Tax, Audit & Advisory',
         'We help businesses with tax filing, audits, compliance, and financial advisory services.'),
        ('Assurance | Taxation | Advisory | Consulting', 'Compliance and Growth Support for Businesses',
         'From statutory compliance to strategic advisory, we support corporates, SMEs, and institutions.'),
        ('Est. 2003 — 20+ Years of Excellence', 'Practical Financial and Regulatory Guidance',
         'Our team delivers clear, partner-led support in tax, audit, corporate law, and finance.'),
    ]
    for i, (tag, title, desc) in enumerate(slides, 1):
        builder.section('HOME', 1, f'Hero Slide {i}', 'Hero', 'Get Consultation | Contact Us')
        builder.el('Tagline', tag, FONT_HERO_TAG)
        builder.el('Title', title, FONT_HERO_TITLE)
        builder.el('Description', desc, FONT_BODY)
    builder.section('HOME', 1, 'Hero trust badges', 'Hero', '')
    for t in ['20+ Years of Experience', 'RBI Registered', 'CAG Empanelled']:
        builder.el('Trust badge', t, FONT_BODY)
    builder.section('HOME', 1, 'Firm highlights', 'Stats row', '')
    for stat in ['20+ Years of Excellence', '7+ Partners', '4 Office Locations', '12+ Service Verticals']:
        builder.el('Stat', stat, FONT_CARD)
    builder.section('HOME', 1, 'SEO intent block', 'Section', '')
    builder.el('Title', 'Chartered Accountant firm for tax, audit, GST, and compliance services in India', FONT_TITLE)
    builder.el('Paragraph', 'Dwivedi Gupta & Co. helps businesses with partner-led tax advisory, statutory and internal audit, GST compliance, company law support, and financial consulting. Explore focused service pages to review scope, deliverables, and ideal fit.', FONT_BODY)
    builder.section('HOME', 1, 'About the Firm', 'Section', 'Learn More About Us | Contact Our Team')
    builder.el('Eyebrow', 'Who We Are', FONT_EYEBROW)
    builder.el('Title', 'About the Firm', FONT_TITLE)
    builder.el('Subtitle', 'Built on trust, guided by expertise, and focused on long-term client value.', FONT_BODY)
    builder.el('Kicker', 'Trusted Since 2003 | Partner-Led Professional Services', FONT_BODY)
    builder.el('Lead title', 'Reliable Tax, Audit, and Advisory Support for Growing Businesses', FONT_CARD)
    for p in [
        'Dwivedi Gupta & Co. delivers practical, compliance-focused solutions with senior partner involvement and consistent execution standards.',
        'Every engagement is tailored to your growth stage, business model, and regulatory needs.',
        'With a strong on-ground office presence and multidisciplinary teams, we deliver responsive support, structured execution, and decision-ready insights for businesses across sectors.',
        'Our teams follow a process-driven approach covering assessment, execution, review, and post-engagement support so clients receive clarity along with consistent implementation quality.',
    ]:
        builder.el('Paragraph', p, FONT_BODY)
    for title, text in [
        ('Credentials', 'RBI Registered and CAG Empanelled.'),
        ('Comprehensive Services', 'Tax, audit, compliance, advisory, and finance support.'),
        ('Client Commitment', 'Long-term relationships built on integrity and responsiveness.'),
    ]:
        builder.el('Highlight card title', title, FONT_CARD)
        builder.el('Highlight card text', text, FONT_BODY)
    services = [
        ('Tax & Regulatory Services', 'Strategic tax planning, compliance, and representation services to optimize tax efficiency while ensuring full regulatory compliance.'),
        ('Audit & Assurance', 'Comprehensive audit solutions including statutory audits, internal audits, tax audits, stock audits, and management audits to strengthen financial controls and transparency.'),
        ('Corporate Law & Compliance', 'Company incorporation, corporate governance support, secretarial services, regulatory filings, and legal compliance assistance.'),
        ('Advisory & Consulting', 'Business advisory services across taxation, regulatory frameworks, LLPs, trusts, and financial regulations with research-based professional insights.'),
        ('Finance & Project Consultancy', 'End-to-end financial consulting including project finance, debt syndication, banking support, and financial structuring for businesses.'),
        ('Government Schemes Consultancy', 'Awareness, planning, and implementation support for Central and State Government schemes across multiple sectors.'),
    ]
    builder.section('HOME', 1, 'Core Services', 'Card grid', 'View All Services')
    builder.el('Eyebrow', 'What We Offer', FONT_EYEBROW)
    builder.el('Title', 'Our Core Tax, Audit, GST, and Advisory Services', FONT_TITLE)
    builder.el('Intro', 'Professional assurance, tax, and advisory solutions designed for modern business requirements.', FONT_BODY)
    for title, desc in services:
        builder.el('Card title', title, FONT_CARD)
        builder.el('Card text', desc, FONT_BODY)
        builder.el('Link', 'Learn more', FONT_BODY)
    industries = [
        ('Manufacturing & Infrastructure', 'Compliance, audit, and financial control frameworks for operationally complex businesses.'),
        ('Banking & Financial Institutions', 'Regulatory-ready support for governance, reporting, and assurance requirements.'),
        ('Real Estate & Construction', 'Project-focused advisory, tax planning, and compliance support across development cycles.'),
        ('Trading & Export Businesses', 'Structured tax and documentation guidance for cross-border operations and growth.'),
        ('Government & Public Sector', 'Audit, compliance, and advisory solutions aligned with institutional accountability standards.'),
        ('SMEs & Startups', 'Scalable financial, tax, and compliance support tailored for fast-moving growth stages.'),
        ('Non-Profit Organizations', 'Transparent reporting and regulatory support for mission-driven organizations.'),
    ]
    builder.section('HOME', 1, 'Industries', 'Carousel', 'Explore Industries')
    builder.el('Eyebrow', 'Sector Expertise', FONT_EYEBROW)
    builder.el('Title', 'Industries We Serve', FONT_TITLE)
    builder.el('Intro', 'Domain-specific advisory, assurance, and compliance solutions built for sector realities.', FONT_BODY)
    for title, desc in industries:
        builder.el('Industry title', title, FONT_CARD)
        builder.el('Industry text', desc, FONT_BODY)
        builder.el('Link', 'View More', FONT_BODY)
    builder.section('HOME', 1, 'Clients preview', 'Section', 'View Our Clients')
    builder.el('Eyebrow', 'Who We Work With', FONT_EYEBROW)
    builder.el('Title', 'Trusted by Leading Organizations', FONT_TITLE)
    builder.el('Intro', 'Corporates, banks, and institutions across India rely on us for their financial and advisory needs.', FONT_BODY)
    builder.el('Caption', 'Representing leading names across banking, manufacturing, and institutions.', FONT_BODY)
    for q, by in [
        ('Their team is highly responsive and always delivers practical guidance on time-sensitive compliance matters.', 'CFO, Manufacturing Group'),
        ('We value their partner-led involvement and clear advisory during audits, tax planning, and strategic decisions.', 'Director, Financial Services Firm'),
        ('A dependable long-term advisor with strong domain knowledge and professional execution standards.', 'Founder, Growth-Stage Enterprise'),
    ]:
        builder.el('Testimonial quote', q, FONT_BODY)
        builder.el('Testimonial attribution', by, FONT_BODY)
    builder.section('HOME', 1, 'CTA band', 'Band', 'Schedule a Consultation')
    builder.el('Eyebrow', 'Expert Guidance Since 2003', FONT_EYEBROW)
    builder.el('Title', 'Looking for Reliable Financial & Advisory Experts?', FONT_TITLE)
    builder.el('Description', 'Partner with Dwivedi Gupta & Co. for professional guidance, compliance confidence, and sustainable business growth.', FONT_BODY)


def build_about(builder):
    builder.section('ABOUT US', 2, 'Hero', 'Page hero', 'Schedule Consultation | Talk to Our Team')
    builder.el('Kicker', 'About Dwivedi Gupta & Co.', FONT_EYEBROW)
    builder.el('Title', 'Partner-led Chartered Accountant advisory, assurance, and compliance expertise since 2003.', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'We help businesses stay compliant, strengthen governance, and build long-term financial confidence.', FONT_BODY)
    builder.section('ABOUT US', 2, 'Who We Are', 'Section', '')
    builder.el('Title', 'Who We Are', FONT_TITLE)
    for p in [
        'Founded in 2003, Dwivedi Gupta & Co. (DGC) is a Chartered Accountants firm delivering taxation, audit, assurance, and advisory services through a strong partner-led approach.',
        'We tailor every engagement to client context, growth stage, and compliance requirements instead of following a one-size-fits-all model.',
        'Our office infrastructure, experienced professionals, and partner-led supervision enable us to handle compliance-intensive assignments for corporates, institutions, and growth-stage businesses with confidence.',
        'We focus on practical outcomes through disciplined planning, timely coordination, and quality review so clients can make decisions with stronger financial and compliance confidence.',
    ]:
        builder.el('Paragraph', p, FONT_BODY)
    for stat in ['2003 Established', '4 Office Locations', '20+ Years of Experience', '7 Partners']:
        builder.el('Stat', stat, FONT_CARD)
    builder.section('ABOUT US', 2, 'Vision and Mission', 'Section', '')
    builder.el('Kicker', 'Our Direction', FONT_EYEBROW)
    builder.el('Title', 'Vision and Mission', FONT_TITLE)
    builder.el('Vision', 'To be a trusted leader in professional financial and advisory services through excellence, innovation, and integrity.', FONT_BODY)
    builder.el('Mission', 'To deliver practical, ethical, and high-impact solutions that strengthen compliance, improve decision-making, and create long-term client value.', FONT_BODY)
    values = [
        ('Partnership', 'We work as an extension of your team with ownership, accountability, and a long-term perspective.'),
        ('Integrity', 'Transparent processes and strong ethical standards guide every advisory and assurance engagement.'),
        ('Passion', 'We are deeply invested in client outcomes and committed to high-quality execution.'),
        ('Excellence', 'Continuous improvement, best practices, and technical depth define our delivery standards.'),
    ]
    builder.section('ABOUT US', 2, 'Core Values', 'Card grid', '')
    builder.el('Kicker', 'What We Stand For', FONT_EYEBROW)
    builder.el('Title', 'Core Values', FONT_TITLE)
    for title, desc in values:
        builder.el('Card title', title, FONT_CARD)
        builder.el('Card text', desc, FONT_BODY)
    builder.section('ABOUT US', 2, 'Team Capability', 'Section', '')
    builder.el('Title', 'Team Capability', FONT_TITLE)
    builder.el('Intro', 'We invest in recruiting and developing highly capable professionals, supported by strong infrastructure and multidisciplinary expertise.', FONT_BODY)
    for m in ['20 Audit Staff', '10 Finance & Consultancy', '8 Tax & Legal', '5 Government Schemes', '7 EDP Operators', '10 Support Staff']:
        builder.el('Metric', m, FONT_CARD)
    builder.section('ABOUT US', 2, 'Our Presence', 'Office grid', '')
    builder.el('Kicker', 'Where We Are', FONT_EYEBROW)
    builder.el('Title', 'Our Presence', FONT_TITLE)
    offices = [
        ('Varanasi', 'Head Office', 'S-8/108-B-3-A Prashantpuri, M.A Road, Varanasi - 221002'),
        ('Delhi', 'Branch Office', '62, Shrestha Vihar, Vikas Marg Extension, Delhi - 110092'),
        ('Kolkata', 'Branch Office', 'Brijdham Housing Complex, 637 Dakshin Dari Road, 5th Floor Flat-5E, Kolkata'),
        ('Bokaro', 'Branch Office', 'C-1, 21A, 2nd Floor, City Centre, Sector-4, Bokaro Steel City, Jharkhand'),
    ]
    for city, kind, addr in offices:
        builder.el('Office', f'{city} — {kind} — {addr}', FONT_BODY)
    builder.section('ABOUT US', 2, 'CTA', 'Band', 'Schedule a Consultation')
    builder.el('Kicker', 'Expert Guidance Since 2003', FONT_EYEBROW)
    builder.el('Title', 'Ready to strengthen compliance and financial decision-making?', FONT_TITLE)
    builder.el('Description', 'Work with our team for partner-led advisory and practical execution support.', FONT_BODY)


def build_services(builder, data):
    builder.section('SERVICES', 3, 'Hero', 'Page hero', 'Schedule Consultation | Speak to Our Team')
    builder.el('Kicker', 'Services Portfolio', FONT_EYEBROW)
    builder.el('Title', 'Partner-led Chartered Accountant services for assurance, taxation, GST, and advisory solutions.', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'We combine deep technical expertise with practical business understanding to deliver reliable, timely, and strategy-aligned outcomes.', FONT_BODY)
    for m in ['20+ Years of practice', '8 Core services', '4 Office locations']:
        builder.el('Metric', m, FONT_CARD)
    builder.section('SERVICES', 3, 'Service cards', 'Card grid', '')
    builder.el('Kicker', 'What We Offer', FONT_EYEBROW)
    builder.el('Title', 'Comprehensive CA and advisory solutions', FONT_TITLE)
    for svc in data['services']:
        builder.el('Card title', svc['title'], FONT_CARD)
        builder.el('Card text', svc['shortDescription'], FONT_BODY)
        for area in svc['keyAreas'][:3]:
            builder.el('Key area', area, FONT_BODY)
        builder.el('Link', 'View Service Details →', FONT_BODY)
    why = [
        ('Partner-Led Engagement', 'Senior partners are directly involved in your engagement, ensuring quality and accountability.'),
        ('Full-Service Under One Roof', 'From incorporation and tax to audit and advisory—we cover the full lifecycle so you have one trusted partner.'),
        ('Timely & Reliable Delivery', 'We meet deadlines and keep you informed so you never miss a compliance date or opportunity.'),
        ('RBI & CAG Empanelled', 'Our empanelment reflects the standards we uphold and our capability to serve institutions.'),
        ('Multi-City Presence', 'Head office in Varanasi with branches in Delhi, Kolkata, and Bokaro for nationwide reach.'),
        ('Industry & Sector Expertise', 'We serve manufacturing, banking, real estate, government, SMEs, and non-profits with tailored approaches.'),
    ]
    builder.section('SERVICES', 3, 'Why DGC', 'Card grid', '')
    builder.el('Title', 'Why businesses choose our service delivery model', FONT_TITLE)
    for title, desc in why:
        builder.el('Card title', title, FONT_CARD)
        builder.el('Card text', desc, FONT_BODY)
    steps = [
        ('01 Consult', 'We understand your business, goals, and challenges through a structured consultation.'),
        ('02 Plan', 'We design a clear scope, timeline, and deliverables aligned with your requirements.'),
        ('03 Deliver', 'Our team executes with quality and keeps you updated at every stage.'),
        ('04 Support', 'Ongoing support and reviews so you stay compliant and informed.'),
    ]
    builder.section('SERVICES', 3, 'Our Approach', 'Steps', '')
    for step, text in steps:
        builder.el('Step', f'{step}: {text}', FONT_BODY)
    builder.section('SERVICES', 3, 'Detailed catalogue', 'List', 'Schedule Consultation | Discuss Requirements')
    builder.el('CTA text', 'Need a custom scope aligned to your business goals? We tailor deliverables, timelines, and engagement structure accordingly.', FONT_BODY)
    for svc in data['services']:
        builder.section('SERVICES', 3, f'Detailed — {svc["title"]}', 'Catalogue row', 'Go to detailed page →')
        builder.el('Long description', svc['longDescription'], FONT_BODY)
        for area in svc['keyAreas']:
            builder.el('Key deliverable', area, FONT_BODY)


def build_service_details(builder, data):
    page_no = 16
    for svc in data['services']:
        builder.section(f'SERVICE DETAIL — {svc["slug"]}', page_no, 'Hero', 'Page hero', 'Schedule Consultation | Talk to Our Team')
        builder.el('Breadcrumb', f'Home / Services / {svc["title"]}', FONT_BODY)
        builder.el('Kicker', 'Service Detail', FONT_EYEBROW)
        builder.el('Title', svc['title'], FONT_PAGE_TITLE)
        builder.el('Subtitle', svc['shortDescription'], FONT_BODY)
        builder.el('Overview', svc['longDescription'], FONT_BODY)
        for snap in [
            'Partner-led supervision and review',
            'Structured timeline and milestone updates',
            'Documentation and compliance-first execution',
            'Cross-functional support for related services',
        ]:
            builder.el('Engagement snapshot', snap, FONT_BODY)
        builder.el('Deliverables heading', f'What we deliver under {svc["title"]}', FONT_TITLE)
        for area in svc['keyAreas']:
            builder.el('Deliverable', area, FONT_BODY)
        builder.el('CTA', 'Schedule Consultation | Contact Us', FONT_BODY)


def build_industries(builder):
    industries = [
        ('Manufacturing & Infrastructure', 'We support manufacturers and infrastructure players with cost accounting, tax incentives, GST compliance, and project finance.',
         ['Tax incentives & exemptions', 'Cost accounting & MIS', 'GST and indirect tax compliance', 'Statutory & internal audit']),
        ('Banking & Financial Institutions', 'RBI compliance, statutory audit, internal audit, and risk assurance for banks, NBFCs, and financial institutions.',
         ['RBI compliance & reporting', 'Statutory & branch audits', 'Internal audit & risk', 'Tax and transfer pricing']),
        ('Real Estate & Construction', 'Project accounting, GST on construction, RERA compliance, and regulatory support for developers and contractors.',
         ['Project & contract accounting', 'GST on works contract', 'RERA and regulatory filings', 'Joint venture structuring']),
        ('Trading & Export Businesses', 'Customs, export benefits, multi-state GST, and inventory accounting for trading and export-oriented businesses.',
         ['Export incentives & MEIS/SEIS', 'Customs and DGFT compliance', 'Multi-state GST registration', 'Inventory & working capital']),
        ('Government & Public Sector', 'CAG empanelment, government audits, and compliance support for public sector undertakings and government schemes.',
         ['CAG & government audits', 'Scheme implementation support', 'Compliance & reporting', 'Procurement and contracts']),
        ('SMEs & Startups', 'Company formation, ESOP accounting, startup tax benefits, and compliance tailored for growing and early-stage businesses.',
         ['Incorporation & secretarial', 'Income tax 80-IAC & incentives', 'ESOP & ESOP valuation', 'GST and annual compliance']),
        ('Non-Profit Organizations', 'Trust and society compliance, 80G/12A support, fund accounting, and audit for NGOs and non-profit entities.',
         ['Trust/society compliance', '80G, 12A & FCRA', 'Fund accounting & audit', 'Board and donor reporting']),
    ]
    builder.section('INDUSTRIES', 4, 'Hero', 'Page hero', 'Explore Services | Book Strategy Call')
    builder.el('Kicker', 'Sector Expertise', FONT_EYEBROW)
    builder.el('Title', 'Industry-specific tax, audit, advisory, and compliance support', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'We align tax, audit, and regulatory execution with your sector obligations, operating model, and strategic priorities.', FONT_BODY)
    for stat in ['7+ Sectors Served', '20+ Years of Experience', '4 Office Locations', '100% Partner-Led Engagement']:
        builder.el('Stat', stat, FONT_CARD)
    for title, desc, services in industries:
        builder.section('INDUSTRIES', 4, title, 'Industry card', 'Discuss Requirements')
        builder.el('Description', desc, FONT_BODY)
        for s in services:
            builder.el('Service bullet', s, FONT_BODY)
    builder.section('INDUSTRIES', 4, 'CTA', 'Band', 'Schedule a Consultation | Contact Us')
    builder.el('Title', 'Your sector strategy starts here', FONT_TITLE)
    builder.el('Description', 'Partner with Dwivedi Gupta & Co. for sector-specific assurance, tax, and advisory support.', FONT_BODY)


def build_team(builder):
    partners = [
        ('CA. Surendra Kumar Dwivedi', 'Founder Partner', 'FCA, FAFP, DISA, B.Com, LLB', 'Direct & Indirect Tax, Assurance & Accounting, Project Finance/Debt Syndication', '33 Years', 'surendra.dwivedi@dgc.ind.in', '+91 9415203012', '074798', ''),
        ('CA. Vivek Anand Mohan', 'Partner', 'FCA, B.S.C., LLB, DISA(ICAI), DIM, CCCAB', 'Direct & Indirect Tax, Assurance & Accounting, Project Finance/Debt Syndication', '18 Years', 'vivek.gupta@dgc.ind.in', '+91 9415804906', '407188', ''),
        ('CA. Aditi Kapoor', 'Partner', 'FCA, DISA (ICAI), DIRM (ICAI), B.Com (H)', 'Financial Advisory, Technical & Economic Viability, Treasury, Direct Tax, Government Schemes', '14 Years', 'aditi.kapoor@dgc.ind.in', '+91 9455113033', '413477', 'In Charge: Delhi Branch'),
        ('CA. Sharad Kumar Jaiswal', 'Partner', 'FCA, B.Com', 'Audit, Project financing, Statutory Compliance', '16 Years', 'sharad.jaiswal@dgc.ind.in', '+91 9026281688', '410050', ''),
        ('CA. Shweta Bharuka', 'Partner', 'ACA, B. Com (H)', 'Direct Tax, Assurance & Accounting', '9 Years', 'swetabharuka@gmail.com', '+91 9831381135', '308394', 'In-charge: Kolkata Branch'),
        ('CA. Neha Nathani', 'Partner', 'FCA and B.Com (H)', 'Taxation and Auditing', '11 Years', 'neha.nathani@dgc.ind.in', '+91 9918115575', '425953', ''),
        ('CA. Breesket Singh', 'Partner', 'FCA and B.Com (H)', 'Taxation and Auditing', '21 Years', 'breesket@gmail.com', '+91 9430309193', '062437', 'In Charge: Bokaro Branch'),
    ]
    builder.section('TEAM', 5, 'Hero', 'Page hero', 'Schedule Consultation | Contact Us')
    builder.el('Title', 'Our Team', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'Experienced partners and professionals delivering trusted financial and advisory guidance.', FONT_BODY)
    for stat in ['7 Partner Profiles', '20+ Years Legacy', 'Pan India Client Service Reach']:
        builder.el('Stat', stat, FONT_CARD)
    for name, role, qual, spec, exp, email, phone, mem, branch in partners:
        builder.section('TEAM', 5, name, 'Partner card', 'View more')
        builder.el('Role', role, FONT_CARD)
        builder.el('Qualifications', qual, FONT_BODY)
        builder.el('Specialization', spec, FONT_BODY)
        builder.el('Experience', exp, FONT_BODY)
        if branch:
            builder.el('Branch', branch, FONT_BODY)
        builder.el('Email', email, FONT_BODY)
        builder.el('Phone', phone, FONT_BODY)
        builder.el('ICAI Membership', mem, FONT_BODY)
    builder.section('TEAM', 5, 'CTA', 'Band', 'Careers | Get in Touch')
    builder.el('Text', 'We recruit and nurture highly capable talent committed to professional excellence.', FONT_BODY)


def build_clients(builder, data):
    builder.section('CLIENTS', 6, 'Hero', 'Page hero', 'Start a Conversation | Schedule Consultation')
    builder.el('Kicker', 'Trusted Partnerships', FONT_EYEBROW)
    builder.el('Title', 'Client relationships built on consistency and trust', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'We support corporates, MSMEs, startups, and institutions with partner-led execution across audit, tax, and compliance functions.', FONT_BODY)
    for logo in data['clients']:
        builder.el('Client logo', logo['name'], FONT_BODY)
    segments = [
        ('Corporates & Listed Companies', 'Statutory audit, tax planning, and compliance for listed and unlisted corporates.'),
        ('MSMEs & Growing Businesses', 'Bookkeeping, GST, and annual compliance tailored for small and medium enterprises.'),
        ('Startups', 'Incorporation, funding compliance, and startup-specific tax and regulatory support.'),
        ('Individuals & HNIs', 'Personal tax planning, ITR filing, and wealth management compliance.'),
    ]
    for title, desc in segments:
        builder.el('Segment title', title, FONT_CARD)
        builder.el('Segment text', desc, FONT_BODY)
    for q, a in [
        ('Professional, timely, and always available. Our go-to firm—DWIVEDI GUPTA & Co.—for the last five years.', 'Client, Manufacturing'),
        ('They simplified our GST and compliance. Highly recommend for startups.', 'Client, Technology'),
    ]:
        builder.el('Testimonial', q, FONT_BODY)
        builder.el('Attribution', a, FONT_BODY)
    for m in ['92%+ Client Retention', '15+ Industries Served', '4 Offices Multi-City Support', '< 24 Hrs Average Response Time']:
        builder.el('Metric', m, FONT_CARD)
    builder.section('CLIENTS', 6, 'CTA', 'Band', 'Start a Conversation')
    builder.el('Text', 'Join our growing list of trusted clients.', FONT_BODY)


def build_insights(builder, data):
    builder.section('INSIGHTS', 7, 'Hero', 'Page hero', 'Schedule Consultation | Contact Us')
    builder.el('Kicker', 'Insights & Knowledge', FONT_EYEBROW)
    builder.el('Title', 'Practical updates for tax, compliance, and advisory decisions', FONT_PAGE_TITLE)
    builder.el('Intro', 'Our insights cover GST, direct and indirect tax, MCA/ROC compliance, audit, startup and MSME schemes, and general advisory. We publish updates and checklists to keep you ahead of due dates and regulatory changes. For advice tailored to your situation, get in touch.', FONT_BODY)
    for cat in ['All', 'Tax', 'GST', 'Compliance', 'Audit', 'Startups', 'Advisory', 'Government Schemes']:
        builder.el('Category filter', cat, FONT_BODY)
    for art in data['insights']:
        builder.section('INSIGHTS', 7, art['title'], 'Article card', 'Read article →')
        builder.el('Category', art['category'], FONT_BODY)
        builder.el('Date', art['dateDisplay'], FONT_BODY)
        builder.el('Summary', art['summary'], FONT_BODY)
        for para in art['body']:
            builder.el('Body paragraph', para, FONT_BODY)
    builder.section('INSIGHTS', 7, 'CTA', 'Band', 'Contact Us | Schedule Consultation')
    builder.el('Text', 'Need guidance on a specific issue? Let us help.', FONT_BODY)


def build_team_members(builder):
    qualified = [
        ('CA Ruchi Singh', 'ACA, B.Com', '3 Years', 'Supports engagement execution with focus on review quality and reporting clarity.'),
        ('CA Viraj Agarwal', 'ACA, B.Com', '3 Years', 'Contributes to audit and compliance assignments with structured documentation support.'),
        ('CA Ajit Dev Pandey', 'ACA, B.Com', '2 Years', 'Handles core working papers and assists in timely delivery of professional assignments.'),
        ('Mr. P M Gupta', 'Diploma in Electrical Engineering, CAIIB (Part I)', 'Retired Banker as Chief Manager, UBI in 2013', 'Brings strong banking operations perspective and credit-process understanding to advisory support.'),
        ('Mr. R. K. Pandey', 'BA, LLB', '26 Years', 'Provides legal and compliance-oriented support for documentation and process alignment.'),
        ('CS Apoorva Singh', 'B.Com (H), ACS', '13 Years', 'Supports secretarial and regulatory workstreams with strong governance-focused execution.'),
        ('CS Urmi Chhaparia', 'B.Com, ACS', '8 Years', 'Assists corporate compliance and filing activities with timeline-driven coordination.'),
        ('Miss Shreya Pandey', 'B.Com, Qualified Company Secretary', 'Newly Qualified', 'Contributes to compliance workflows with updated regulatory knowledge and diligence.'),
    ]
    semi = [
        ('Shruti Khemka', 'CA Intermediate (Pursuing CA Final)', 'Supports documentation and preliminary working schedules for assignment teams.'),
        ('Ritika Khosala', 'CA Intermediate (Pursuing CA Final)', 'Assists in compliance data preparation and file organization across engagements.'),
        ('Jasraj Singh', 'CA Intermediate (Pursuing CA Final)', 'Contributes to audit support activities and checklist-based execution tasks.'),
        ('Shivam Agrwal', 'CA Intermediate (Pursuing CA Final)', 'Works on backend schedules and reconciliation support for team deliverables.'),
        ('Vishesh Misra', 'CA Intermediate (Pursuing CA Final)', 'Provides process support for periodic filings and assignment documentation.'),
        ('Harsh Jaiswal', 'CA Intermediate (Pursuing CA Final)', 'Supports quality checks and structured working-paper updates for execution teams.'),
        ('Nehal Ahmad', 'CA Intermediate (Pursuing CA Final)', 'Assists engagement teams with data collation and compliance task tracking.'),
        ('Vaibhav Jaiswal', 'CA Intermediate (Pursuing CA Final)', 'Contributes to audit and compliance support under supervised workflows.'),
        ('Shreya Gupta', 'CA Intermediate (Pursuing CA Final)', 'Supports assignment coordination and readiness of key records and checklists.'),
        ('Yash Agrawal', 'CS Executive (Pursuing CS Professional)', 'Assists corporate compliance workstreams and secretarial process documentation.'),
        ('Kajal Parikh', 'CS Executive (Pursuing CS Professional)', 'Supports filing activities and regulatory documentation with process discipline.'),
    ]
    builder.section('TEAM MEMBERS', 18, 'Hero', 'Page hero', 'View Partners | Contact Our Team')
    builder.el('Kicker', 'Team Members', FONT_EYEBROW)
    builder.el('Title', 'Our People', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'We recruit, train, motivate and retain highly capable and sharp talent who bring quality to their work and deliver the best solutions. We nurture our people and turn them into our strongest assets.', FONT_BODY)
    builder.section('TEAM MEMBERS', 18, 'Qualified Staff', 'Card grid', '')
    builder.el('Section title', 'Experienced professionals supporting engagement delivery', FONT_TITLE)
    for name, qual, exp, line in qualified:
        builder.el('Member name', name, FONT_CARD)
        builder.el('Qualification', qual, FONT_BODY)
        builder.el('Experience', exp, FONT_BODY)
        builder.el('Description', line, FONT_BODY)
    builder.section('TEAM MEMBERS', 18, 'Semi-Qualified Staff', 'Card grid', 'Schedule Consultation | Discuss Requirement')
    for name, qual, line in semi:
        builder.el('Member name', name, FONT_CARD)
        builder.el('Qualification', qual, FONT_BODY)
        builder.el('Description', line, FONT_BODY)


def build_support_team(builder):
    builder.section('SUPPORT TEAM', 19, 'Hero', 'Page hero', 'Back to Partners | View Team Members')
    builder.el('Kicker', 'Support Team', FONT_EYEBROW)
    builder.el('Title', 'Support Team Details', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'We will publish the complete Support Team structure and profiles here once you share the final content.', FONT_BODY)
    builder.el('Section title', 'Dedicated manpower across audit, advisory, tax, and execution support', FONT_TITLE)
    for m in ['20 Audit Staff', '10 Finance & Consultancy', '08 Tax & Legal Section', '05 Government Schemes & Consultancy', '7 E.D.P. Operators', '10 Semi-skilled Staff/Peons']:
        builder.el('Metric', m, FONT_CARD)
    builder.section('SUPPORT TEAM', 19, 'CTA', 'Band', 'Contact Us | Schedule Consultation')


def build_other_pages(builder):
    builder.section('CAREERS', 8, 'Main', 'Page hero + text', 'Contact Us')
    builder.el('Kicker', 'Work With Us', FONT_EYEBROW)
    builder.el('Title', 'Careers', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'Join our team of professionals. Explore opportunities at Dwivedi Gupta & Co.', FONT_BODY)
    builder.el('Body', 'We are always looking for talented chartered accountants and finance professionals. Email shikhar.dwivedi@dgc.ind.in', FONT_BODY)

    builder.section('SCHEDULE CONSULTATION', 9, 'Hero', 'sched-hero', 'Explore Services | Talk to Our Team')
    builder.el('Kicker', 'Schedule Consultation', FONT_EYEBROW)
    builder.el('Title', 'Book Your Consultation Session', FONT_PAGE_TITLE)
    builder.el('Description', 'Reserve a focused 30-minute advisory call with our CA team. Share your preferred slot and service area, and we will confirm the engagement mode within 1-2 business days.', FONT_BODY)
    for t in ['Free 30-minute discussion', 'No obligation advisory', 'Confidential interaction']:
        builder.el('Trust chip', t, FONT_BODY)
    for title, desc in [
        ('Free 30-minute discussion', 'Understand your needs with no cost or obligation—get clarity before you commit.'),
        ('Your choice of format', 'In-person at our office, phone call, or video call—we adapt to your preference.'),
        ('Partner-led expertise', 'Speak directly with experienced CAs who can advise on tax, audit, and compliance.'),
    ]:
        builder.el('Why book title', title, FONT_CARD)
        builder.el('Why book text', desc, FONT_BODY)
    for label in ['Full name', 'Email', 'Phone', 'Service of interest', 'Preferred date', 'Preferred time', 'Additional details']:
        builder.el('Form label', label, FONT_BODY)
    for opt in ['Tax Planning & Compliance', 'Audit & Assurance', 'Accounting & Bookkeeping', 'GST Advisory', 'Company Formation & Compliance', 'Financial Advisory', 'Other']:
        builder.el('Service option', opt, FONT_BODY)
    builder.el('Button', 'Request Consultation', FONT_BODY)
    builder.el('Success message', 'Your consultation request has been received. We will confirm your slot shortly.', FONT_BODY)

    builder.section('CONTACT', 10, 'Hero', 'contact-hero', 'Schedule Consultation | Chat on WhatsApp')
    builder.el('Kicker', 'Get in Touch', FONT_EYEBROW)
    builder.el('Title', 'Contact Our CA Experts', FONT_PAGE_TITLE)
    builder.el('Intro', 'Whether you need tax advice, audit support, GST compliance, or company formation—our team is here to help. Reach out via the form, email, or phone. We respond to all enquiries within 24–48 hours.', FONT_BODY)
    builder.section('CONTACT', 10, 'Reach Us Directly', 'Card', '')
    builder.el('Email', 'shikhar.dwivedi@dgc.ind.in', FONT_BODY)
    builder.el('Mobile', '+91 9721227799', FONT_BODY)
    builder.el('Landline', '0542-2502525', FONT_BODY)
    builder.el('WhatsApp', '+91 9721227799', FONT_BODY)
    for item in [
        'Response within 24–48 hours on working days',
        'Confidentiality on all enquiries',
        'Support in Tax, Audit, GST, Company Formation, and more',
        'Option to schedule a consultation for detailed discussions',
        'Clear checklist of documents and next steps after first response',
        'Partner-led review for complex compliance and advisory matters',
        'Transparent communication on timeline and scope before working on the matter.',
    ]:
        builder.el('What to expect', item, FONT_BODY)
    for label in ['Name', 'Email', 'Phone', 'Subject', 'Message']:
        builder.el('Form label', label, FONT_BODY)
    for subj in ['General Enquiry', 'Tax', 'Audit', 'GST', 'Company Formation', 'Compliance', 'Other']:
        builder.el('Subject option', subj, FONT_BODY)
    builder.el('Button', 'Send Message', FONT_BODY)
    builder.el('Success message', 'Thank you. We will get back to you within 24–48 hours.', FONT_BODY)
    offices = [
        ('Head Office — Varanasi', 'S-8/108-B-3-A Prashantpuri, M.A Road, Varanasi – 221002', 'Our flagship office for assurance, tax, and advisory. Visit by appointment.'),
        ('Branch Office — Delhi', '62, Shrestha Vihar, Vikas Marg Extension, Delhi – 110092', 'Serving corporates and institutions in the capital region.'),
        ('Branch Office — Kolkata', 'Brijdham Housing Complex, 637 Dakshin Dari Road, 5th Floor Flat-5E, Building No 16-C, Kolkata, West Bengal', 'Eastern India presence for audit, tax, and compliance.'),
        ('Branch Office — Bokaro', 'C-1, 21A, 2nd Floor, City Centre, Sector-4, Bokaro Steel City, Jharkhand', 'Industry-focused CA services in the region.'),
    ]
    for label, addr, desc in offices:
        builder.el('Office label', label, FONT_CARD)
        builder.el('Address', addr, FONT_BODY)
        builder.el('Description', desc, FONT_BODY)
        builder.el('Link', 'Get directions →', FONT_BODY)

    builder.section('COMPLIANCE', 11, 'Hero', 'Page hero', '')
    builder.el('Kicker', 'Regulatory Profile', FONT_EYEBROW)
    builder.el('Title', 'Compliance Information', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'Core statutory and registration details of Dwivedi Gupta & Co. for reference and verification support.', FONT_BODY)
    for label, val in [
        ('Year of Establishment', '2003'),
        ('Firm Registration No.', '012584C'),
        ('Category of Firm (ICAI)', 'I (One)'),
        ('Official Email', 'aditi.kapoor@dgc.ind.in'),
        ('Head Office', 'Varanasi'),
        ('Branch Offices', 'Kolkata, Delhi & Bokaro'),
    ]:
        builder.el(label, val, FONT_BODY)
    builder.el('Note', 'Last updated: 2026. Information is subject to updates as per regulatory changes.', FONT_BODY)

    builder.section('PRIVACY POLICY', 12, 'Hero', 'Page hero', '')
    builder.el('Title', 'Privacy Policy', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'How Dwivedi Gupta & Co. collects, uses, stores, and safeguards information shared through our website and services.', FONT_BODY)
    builder.el('Commitment', 'We are committed to responsible data handling aligned with professional ethics, confidentiality obligations, and applicable data protection laws.', FONT_BODY)
    privacy_sections = {
        'Information We Collect': [
            'Basic contact details such as name, email address, phone number, and organization details when you submit forms.',
            'Engagement information you voluntarily share regarding accounting, taxation, audit, or advisory requirements.',
            'Technical website usage data such as browser type, device details, pages visited, and approximate location for analytics.',
        ],
        'How We Use Your Information': [
            'To respond to consultation requests, service inquiries, and support communications.',
            'To evaluate requirements, prepare proposals, and deliver professional services under agreed scope.',
            'To improve website performance, content relevance, and client communication quality.',
        ],
        'Data Protection & Security': [
            'We apply administrative, technical, and organizational safeguards to protect personal information from unauthorized access.',
            'Access to personal information is restricted to authorized team members and service partners on a need-to-know basis.',
            'Information retention periods are aligned with legal, regulatory, and professional record-keeping requirements.',
        ],
        'Your Rights': [
            'You may request access, correction, or deletion of personal information where legally permissible.',
            'You may withdraw consent for non-essential communication at any time by contacting us directly.',
            'For privacy requests, email our support team with subject line: "Privacy Request".',
        ],
    }
    for title, points in privacy_sections.items():
        builder.el('Section title', title, FONT_CARD)
        for p in points:
            builder.el('Bullet', p, FONT_BODY)
    builder.el('Footer note', 'Last updated: April 2026. Policy content may be revised to reflect legal or operational updates.', FONT_BODY)

    builder.section('TERMS', 13, 'Hero', 'Page hero', '')
    builder.el('Title', 'Terms & Conditions', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'Terms governing use of this website, informational content, and interactions with Dwivedi Gupta & Co.', FONT_BODY)

    builder.section('DISCLAIMER', 14, 'Hero', 'Page hero', '')
    builder.el('Title', 'Disclaimer', FONT_PAGE_TITLE)
    builder.el('Intro', 'This disclaimer is intended to set clear expectations around the use of website materials and communication.', FONT_BODY)
    for title, text in [
        ('Informational Purpose Only', 'Content published on this website is intended for general informational use and should not be interpreted as professional tax, legal, audit, or financial advice for any specific case.'),
        ('No Solicitation', 'Nothing on this website is intended to solicit clients in violation of professional standards.'),
        ('Accuracy of Information', 'We endeavor to keep information accurate and current; however, we do not warrant completeness, reliability, or timeliness of all content at all times.'),
        ('Professional Advice Requirement', 'Before acting on any information from this website, users should obtain direct professional advice based on their facts, records, and applicable regulations.'),
        ('Limitation of Liability', 'Dwivedi Gupta & Co., its partners, and team members are not responsible for any loss arising from reliance on website content without formal professional consultation.'),
    ]:
        builder.el('Card title', title, FONT_CARD)
        builder.el('Card text', text, FONT_BODY)
    builder.el('Updated', 'Last updated: April 2026.', FONT_BODY)

    builder.section('SITEMAP', 15, 'Hero', 'Page hero', '')
    builder.el('Title', 'Sitemap', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'Explore all public pages on the Dwivedi Gupta & Co. website in one structured view.', FONT_BODY)
    builder.el('Note', 'Some pages may be updated over time as we refresh services, articles, and regulatory content.', FONT_BODY)


def build_footer(builder):
    builder.section('FOOTER (all pages)', 17, 'Brand column', 'Footer', 'Schedule Consultation')
    builder.el('Company name', 'Dwivedi Gupta & Co.', FONT_CARD)
    builder.el('Tagline', 'Tax, audit, advisory, and financial consulting since 2003.', FONT_BODY)
    builder.el('Tagline', 'Serving businesses with partner-led expertise, compliance clarity, and long-term value.', FONT_BODY)
    builder.el('Email', 'shikhar.dwivedi@dgc.ind.in', FONT_BODY)
    builder.el('Mobile', '+91 9721227799', FONT_BODY)
    builder.el('Landline', '0542-2502525', FONT_BODY)
    builder.el('Social', 'LinkedIn | mailto:shikhar.dwivedi@dgc.ind.in | tel:+919721227799 | tel:+915422502525 | WhatsApp float (+91 9721227799)', FONT_BODY)
    builder.el('Newsletter heading', 'Newsletter', FONT_CARD)
    builder.el('Placeholder', 'Your email', FONT_BODY)
    builder.el('Button', 'Subscribe', FONT_BODY)
    builder.el('Success', 'Thank you for subscribing.', FONT_BODY)
    builder.el('Note', 'Monthly updates on tax, compliance, and regulatory changes.', FONT_BODY)
    builder.el('Copyright', '© {year} Dwivedi Gupta & Co. All Rights Reserved.', FONT_BODY)
    builder.el('WhatsApp label', 'Chat with us on WhatsApp', FONT_BODY)


def write_workbook(builder, data):
    wb = openpyxl.Workbook()
    # Contact Data sheet
    cds = wb.active
    cds.title = 'Contact Data'
    contact_rows = [
        ['Section', 'Field', 'Value', 'Notes', 'Last Updated'],
        ['Contact', 'Primary Email', 'shikhar.dwivedi@dgc.ind.in', 'Site-wide', str(date.today())],
        ['Contact', 'Mobile', '+91 9721227799', 'Contact page & footer', str(date.today())],
        ['Contact', 'Landline', '0542-2502525', 'Contact page & footer', str(date.today())],
        ['Contact', 'WhatsApp', '+91 9721227799', 'wa.me/919721227799', str(date.today())],
    ]
    for r, row in enumerate(contact_rows, 1):
        for c, val in enumerate(row, 1):
            cds.cell(row=r, column=c, value=val)

    ws = wb.create_sheet('Content Audit')
    headers = ['S.no', 'Page Name', 'Page No', 'Page Sub-heading', 'Design Used', 'Buttons', 'Content', None, None, 'Remarks']
    subheaders = [None, None, None, None, None, None, 'Heading', 'Font name & Size', 'Text', None]
    for c, val in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=val)
    for c, val in enumerate(subheaders, 1):
        ws.cell(row=2, column=c, value=val)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    for cell in ws[1] + ws[2]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill

    for r, row in enumerate(builder.rows, 3):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 32
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 42
    ws.column_dimensions['I'].width = 80
    ws.column_dimensions['J'].width = 36
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    meta = wb.create_sheet('Audit Meta')
    meta.append(['Generated', str(date.today())])
    meta.append(['Source', 'client/src frontend (live user-visible text)'])
    meta.append(['Total content rows', len(builder.rows)])
    meta.append(['Services', len(data['services'])])
    meta.append(['Insights', len(data['insights'])])

    return wb


def main():
    data = load_data()
    builder = AuditBuilder()
    build_header(builder)
    build_home(builder)
    build_about(builder)
    build_services(builder, data)
    build_industries(builder)
    build_team(builder)
    build_clients(builder, data)
    build_insights(builder, data)
    build_other_pages(builder)
    build_service_details(builder, data)
    build_team_members(builder)
    build_support_team(builder)
    build_footer(builder)

    wb = write_workbook(builder, data)
    for path in OUT_PATHS:
        wb.save(path)
        print(f'Saved {path} ({len(builder.rows)} content rows)')


if __name__ == '__main__':
    main()
