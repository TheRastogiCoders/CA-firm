"""
Generate WEBSITE_CONTENT_ON_SCREEN.xlsx — user-visible website content
in the same layout as WEBSITE_CONTENT_AUDIT.xlsx (does not overwrite the old audit).
"""
import json
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
DATA_JSON = Path(__file__).resolve().parent / 'audit_data.json'
OUT_PATH = ROOT / 'WEBSITE_CONTENT_ON_SCREEN.xlsx'
OUT_PATH_FALLBACK = ROOT / 'WEBSITE_CONTENT_ON_SCREEN_BY_PAGE.xlsx'

FONT_BODY = 'DM Sans, 400, ~1rem (body)'
FONT_EYEBROW = 'DM Sans, 600, ~0.8125rem (section eyebrow)'
FONT_TITLE = 'DM Sans, 600, clamp(1.5–2.25rem) (.about-title / .services-title)'
FONT_HERO_TITLE = 'DM Sans, 600, clamp(1.95–3.25rem) (.hero-title)'
FONT_HERO_TAG = 'DM Sans, 500–600, clamp(0.92–1rem) (.hero-tagline)'
FONT_PAGE_TITLE = 'DM Sans, 600, clamp(1.9–3rem) (.page-title)'
FONT_CARD = 'DM Sans, 600, ~1.0625–1.25rem (cards)'
REMARKS = 'Auto-generated from live frontend (Jul 2026). Includes Services mega-menu. Typography: DM Sans.'


# Excel sheet tab names (max 31 chars) for each logical page key
SHEET_NAME_MAP = {
    'HEADER (all pages)': '01 Header',
    'HOME': '02 Home',
    'ABOUT US': '03 About Us',
    'SERVICES': '04 Services',
    'INDUSTRIES': '05 Industries',
    'TEAM': '06 Team',
    'CLIENTS': '07 Clients',
    'INSIGHTS': '08 Insights',
    'CAREERS': '09 Careers',
    'SCHEDULE CONSULTATION': '10 Schedule Consult',
    'CONTACT': '11 Contact',
    'COMPLIANCE': '12 Compliance',
    'PRIVACY POLICY': '13 Privacy Policy',
    'TERMS': '14 Terms',
    'DISCLAIMER': '15 Disclaimer',
    'SITEMAP': '16 Sitemap',
    'FOOTER (all pages)': '17 Footer',
    'TEAM MEMBERS': '18 Team Members',
    'SUPPORT TEAM': '19 Support Team',
    'SHARED CTA BAND (defaults)': '20 Shared CTA Band',
}

# Matches client Header.jsx SERVICES_MEGA_COLUMNS (titles + first 4 keyAreas shown)
SERVICES_MEGA_COLUMNS = [
    ['tax-compliance', 'gst-advisory'],
    ['audit-assurance', 'financial-consulting', 'project-finance'],
    ['corporate-law', 'company-formation', 'government-schemes-advisory'],
]


def sheet_name_for(page_name: str) -> str:
    if page_name in SHEET_NAME_MAP:
        return SHEET_NAME_MAP[page_name]
    if page_name.startswith('SERVICE DETAIL'):
        slug = page_name.replace('SERVICE DETAIL — ', '').replace('SERVICE DETAIL - ', '')
        short = slug.replace('-', ' ').title()
        name = f'Svc {short}'
        return name[:31]
    return page_name[:31]


class AuditBuilder:
    def __init__(self):
        self.rows = []
        self.sno = 0
        self._page = None
        self._sec = None
        self.page_order = []  # stable order of page keys as content is added

    def add(self, page_name, page_no, subheading, design, buttons, element, text, font=FONT_BODY, remarks=None):
        self.sno += 1
        if page_name and page_name not in self.page_order:
            self.page_order.append(page_name)
        self._page = page_name or self._page
        self.rows.append(
            {
                'sno': self.sno,
                'page_name': self._page,
                'page_no': page_no,
                'subheading': subheading,
                'design': design,
                'buttons': buttons or None,
                'element': element,
                'font': font,
                'text': text,
                'remarks': remarks or REMARKS,
            }
        )

    def section(self, page, page_no, sub, design='', buttons=''):
        self._sec = (page, page_no, sub, design, buttons)

    def el(self, element, text, font=FONT_BODY, buttons=None):
        p, pn, sub, design, btns = self._sec
        self.add(p, pn, sub, design, buttons or btns, element, text, font, None)


def load_data():
    with open(DATA_JSON, encoding='utf-8') as f:
        return json.load(f)


def build_header(builder, data):
    builder.section('HEADER (all pages)', 0, 'Brand', 'Header', '')
    builder.el('Firm name', 'DWIVEDI GUPTA & Co.', FONT_CARD)
    builder.el('Tagline', 'Chartered Accountants', FONT_BODY)
    builder.section('HEADER (all pages)', 0, 'Navigation', 'Nav', '')
    for link in ['Home', 'About Us', 'Services', 'Industries', 'Team', 'Clients', 'Insights', 'Contact']:
        builder.el('Nav link', link, FONT_BODY)

    by_slug = {s['slug']: s for s in data.get('services', [])}
    builder.section(
        'HEADER (all pages)',
        0,
        'Services mega-menu',
        'Glass mega panel (3 columns)',
        'All Services ->',
    )
    builder.el('Panel label', 'Our Services', FONT_EYEBROW)
    builder.el('Top link', 'All Services ->', FONT_BODY)
    builder.el(
        'Layout note',
        'Desktop: compact semi-transparent navy glass panel under Services. Mobile: stacked accordion. Up to 4 key areas per category.',
        FONT_BODY,
    )
    for col_index, slugs in enumerate(SERVICES_MEGA_COLUMNS, 1):
        for slug in slugs:
            svc = by_slug.get(slug)
            if not svc:
                continue
            builder.section(
                'HEADER (all pages)',
                0,
                f'Services mega — Column {col_index} — {svc["title"]}',
                'Mega category',
                '',
            )
            builder.el('Category title', svc['title'], FONT_CARD)
            for area in (svc.get('keyAreas') or [])[:4]:
                builder.el('Key area link', area, FONT_BODY)

    builder.section('HEADER (all pages)', 0, 'Team dropdown', 'Dropdown', '')
    for t in ['Partners', 'Team Members', 'Support Team']:
        builder.el('Menu item', t, FONT_BODY)


def build_home(builder):
    builder.section('HOME', 1, 'Hero', 'Hero', 'Schedule Consultation | Contact Us')
    builder.el('Tagline', 'Assurance | Taxation | Advisory | Consulting', FONT_HERO_TAG)
    builder.el('Title', 'Compliance and Growth Support for Businesses', FONT_HERO_TITLE)
    builder.el('Description', 'Tax, audit, and advisory support for corporates and SMEs.', FONT_BODY)

    builder.section('HOME', 1, 'Firm highlights', 'Stats row', '')
    for stat in ['20+ Years of Excellence', '7+ Partners', '4 Office Locations', '12+ Service Verticals']:
        builder.el('Stat', stat, FONT_CARD)

    builder.section('HOME', 1, 'About the Firm', 'Section', 'About the Firm')
    builder.el('Eyebrow', 'Who We Are', FONT_EYEBROW)
    builder.el('Title', 'Dwivedi Gupta & Co.', FONT_TITLE)
    builder.el(
        'Paragraph',
        'A Chartered Accountants firm established in 2003, based in Varanasi with offices in Delhi, Kolkata, and Bokaro. We provide tax, audit, GST, company law, and advisory support for businesses and institutions.',
        FONT_BODY,
    )

    builder.section('HOME', 1, 'Core Services', 'Card grid', 'View All Services')
    builder.el('Eyebrow', 'What We Offer', FONT_EYEBROW)
    builder.el('Title', 'Our Core Services', FONT_TITLE)
    builder.el('Intro', 'Tax, audit, company law, and advisory for businesses and institutions.', FONT_BODY)
    for title, desc in [
        ('Tax & Regulatory', 'Tax planning, compliance, and representation.'),
        ('Audit & Assurance', 'Statutory, internal, tax, and stock audits.'),
        ('Corporate Law', 'Company law, governance, and ROC filings.'),
        ('Advisory & Consulting', 'Valuation, due diligence, and transactions.'),
    ]:
        builder.el('Card title', title, FONT_CARD)
        builder.el('Card text', desc, FONT_BODY)
        builder.el('Link', 'Learn more', FONT_BODY)

    builder.section('HOME', 1, 'Testimonials', 'Card grid', '')
    builder.el('Eyebrow', 'Testimonials', FONT_EYEBROW)
    builder.el('Title', 'What Clients Say', FONT_TITLE)
    for q, by in [
        ('Responsive team with practical guidance on urgent compliance matters.', 'Kajaria Tiles'),
        ('Clear advice on audits, tax planning, and regulatory questions.', "Haldiram's"),
        ('Dependable advisor with sound technical knowledge and timely delivery.', 'RC Rungta Group'),
    ]:
        builder.el('Testimonial quote', q, FONT_BODY)
        builder.el('Testimonial attribution', by, FONT_BODY)

    builder.section('HOME', 1, 'CTA band', 'Band', 'Schedule Consultation | Contact Us')
    builder.el('Title', 'Ready to discuss your requirements?', FONT_TITLE)
    builder.el(
        'Description',
        'Book a consultation or send us a query. We respond within 24 to 48 working hours.',
        FONT_BODY,
    )


def build_about(builder):
    builder.section('ABOUT US', 2, 'Hero', 'Page hero', 'Schedule Consultation | Contact Us')
    builder.el('Kicker', 'About Us', FONT_EYEBROW)
    builder.el('Title', 'Dwivedi Gupta & Co.', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'Chartered Accountants for tax, audit, and advisory since 2003.', FONT_BODY)

    builder.section('ABOUT US', 2, 'Who We Are', 'Section', 'Meet Our Team | Our Services')
    builder.el('Title', 'Who We Are', FONT_TITLE)
    for p in [
        'Dwivedi Gupta & Co. is a Chartered Accountants firm based in Varanasi, with offices in Delhi, Kolkata, and Bokaro. We help businesses with tax, audit, GST, company law, and financial advisory.',
        'Partners stay involved in significant assignments so you get practical advice and clear timelines.',
    ]:
        builder.el('Paragraph', p, FONT_BODY)
    for stat in ['2003 Established', '20+ Years', '4 Offices', '7 Partners']:
        builder.el('Stat', stat, FONT_CARD)

    builder.section('ABOUT US', 2, 'Vision and Mission', 'Section', '')
    builder.el('Vision', 'To be a trusted Chartered Accountants firm known for sound advice and consistent professional service.', FONT_BODY)
    builder.el('Mission', 'To provide practical tax, audit, and advisory support that helps clients stay compliant and manage financial matters with confidence.', FONT_BODY)

    builder.section('ABOUT US', 2, 'Our Offices', 'Office grid', '')
    for office in [
        'Varanasi (Head Office)',
        'Delhi (Branch)',
        'Kolkata (Branch)',
        'Bokaro (Branch)',
    ]:
        builder.el('Office', office, FONT_BODY)

    builder.section('ABOUT US', 2, 'CTA', 'Band', 'Schedule Consultation | Contact Us')
    builder.el('Title', 'Need help with tax, audit, or compliance?', FONT_TITLE)
    builder.el(
        'Description',
        'Speak with our team about your requirements. We respond within 24 to 48 working hours.',
        FONT_BODY,
    )


def build_services(builder, data):
    builder.section('SERVICES', 3, 'Hero', 'Page hero', 'Schedule Consultation | Contact Us')
    builder.el('Kicker', 'Services', FONT_EYEBROW)
    builder.el('Title', 'Tax, Audit, GST & Advisory', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'Clear scope and deliverables for businesses and institutions.', FONT_BODY)

    builder.section('SERVICES', 3, 'Service cards', 'Card grid', '')
    builder.el('Title', 'Our Services', FONT_TITLE)
    for svc in data['services']:
        builder.el('Card title', svc['title'], FONT_CARD)
        builder.el('Card text', svc['shortDescription'], FONT_BODY)
        builder.el('Link', 'Learn more →', FONT_BODY)

    builder.section('SERVICES', 3, 'CTA', 'Band', 'Schedule Consultation | Contact Us')
    builder.el('Title', 'Not sure which service you need?', FONT_TITLE)
    builder.el(
        'Description',
        'Tell us about your requirement. We will suggest the right scope and next steps.',
        FONT_BODY,
    )

    for svc in data['services']:
        builder.section('SERVICES', 3, f'Catalogue — {svc["title"]}', 'Detail excerpt', '')
        builder.el('Long description', svc['longDescription'], FONT_BODY)
        for area in svc['keyAreas']:
            builder.el('Key area', area, FONT_BODY)


def build_service_details(builder, data):
    page_no = 16
    for svc in data['services']:
        builder.section(f'SERVICE DETAIL — {svc["slug"]}', page_no, 'Hero', 'Page hero', 'Schedule Consultation | Contact Us')
        builder.el('Breadcrumb', f'Services / {svc["title"]}', FONT_BODY)
        builder.el('Title', svc['title'], FONT_PAGE_TITLE)
        builder.el('Subtitle', svc['shortDescription'], FONT_BODY)
        builder.el('Overview heading', 'Overview', FONT_TITLE)
        builder.el('Overview', svc['longDescription'], FONT_BODY)
        builder.el('Deliverables heading', 'Key Areas', FONT_TITLE)
        for area in svc['keyAreas']:
            builder.el('Deliverable', area, FONT_BODY)
        builder.el('Related heading', 'Related Services', FONT_TITLE)
        builder.el('CTA title', f'Discuss {svc["title"]}', FONT_TITLE)
        builder.el(
            'CTA description',
            'Share your requirements. We will confirm scope and timelines before starting.',
            FONT_BODY,
        )


def build_industries(builder):
    industries = [
        'Manufacturing & Infrastructure',
        'Banking & Financial Institutions',
        'Real Estate & Construction',
        'Trading & Export',
        'Government & Public Sector',
        'SMEs & Startups',
        'Non-Profit Organizations',
    ]
    builder.section('INDUSTRIES', 4, 'Hero', 'Page hero', 'View Services | Schedule Consultation')
    builder.el('Kicker', 'Industries', FONT_EYEBROW)
    builder.el('Title', 'Sectors We Serve', FONT_PAGE_TITLE)
    builder.el(
        'Subtitle',
        'Tax, audit, and compliance support shaped by your industry rules and operating model.',
        FONT_BODY,
    )
    builder.section('INDUSTRIES', 4, 'List', 'Industry cards', '')
    builder.el('Title', 'Our Industry Focus', FONT_TITLE)
    for title in industries:
        builder.el('Industry title', title, FONT_CARD)
    builder.section('INDUSTRIES', 4, 'CTA', 'Band', 'Schedule Consultation | Contact Us')
    builder.el('Title', 'Need support in your sector?', FONT_TITLE)
    builder.el(
        'Description',
        'Tell us about your industry and requirements. We will outline the right scope and next steps.',
        FONT_BODY,
    )


def build_team(builder):
    partners = [
        ('CA. Surendra Kumar Dwivedi', 'Founder Partner', 'FCA, FAFP, DISA, B.Com, LLB',
         'Direct & Indirect Tax, Assurance & Accounting, Project Finance/Debt Syndication',
         '33 Years', 'surendra.dwivedi@dgc.ind.in', '+91 9415203012', '074798', ''),
        ('CA. Vivek Anand Mohan', 'Partner', 'FCA, B.S.C., LLB, DISA (ICAI), DIM, CCCAB',
         'Direct & Indirect Tax, Assurance & Accounting, Project Finance/Debt Syndication',
         '18 Years', 'vivek.gupta@dgc.ind.in', '+91 9415804906', '407188', ''),
        ('CA. Aditi Kapoor', 'Partner', 'FCA, DISA (ICAI), DIRM (ICAI), B.Com (H)',
         'Financial Advisory, Technical & Economic Viability, Treasury, Direct Tax, Government Schemes',
         '14 Years', 'aditi.kapoor@dgc.ind.in', '+91 9455113033', '413477', 'In Charge: Delhi Branch'),
        ('CA. Sharad Kumar Jaiswal', 'Partner', 'FCA, B.Com',
         'Audit, Project financing, Statutory Compliance',
         '16 Years', 'sharad.jaiswal@dgc.ind.in', '+91 9026281688', '410050', ''),
        ('CA. Shweta Bharuka', 'Partner', 'ACA, B. Com (H)',
         'Direct Tax, Assurance & Accounting',
         '9 Years', 'swetabharuka@gmail.com', '+91 9831381135', '308394', 'In-charge: Kolkata Branch'),
        ('CA. Neha Nathani', 'Partner', 'FCA and B.Com (H)',
         'Taxation and Auditing',
         '11 Years', 'neha.nathani@dgc.ind.in', '+91 9918115575', '425953', ''),
        ('CA. Breesket Singh', 'Partner', 'FCA and B.Com (H)',
         'Taxation and Auditing',
         '21 Years', 'breesket@gmail.com', '+91 9430309193', '062437', 'In Charge: Bokaro Branch'),
    ]
    builder.section('TEAM', 5, 'Hero', 'Page hero', 'Team Members | Contact Us')
    builder.el('Kicker', 'Team', FONT_EYEBROW)
    builder.el('Title', 'Our Partners', FONT_PAGE_TITLE)
    builder.el(
        'Subtitle',
        'Seven partners leading tax, audit, compliance, and advisory across four offices.',
        FONT_BODY,
    )
    for name, role, qual, spec, exp, email, phone, mem, branch in partners:
        builder.section('TEAM', 5, name, 'Partner card', 'View profile')
        builder.el('Role', role, FONT_CARD)
        builder.el('Qualifications', qual, FONT_BODY)
        builder.el('Specialization', spec, FONT_BODY)
        builder.el('Experience', exp, FONT_BODY)
        if branch:
            builder.el('Branch', branch, FONT_BODY)
        builder.el('Email', email, FONT_BODY)
        builder.el('Phone', phone, FONT_BODY)
        builder.el('ICAI Membership', mem, FONT_BODY)

    builder.section('TEAM', 5, 'CTA', 'Band', 'Contact Us | Careers')
    builder.el('Title', 'Want to work with our team?', FONT_TITLE)
    builder.el('Description', 'Reach out for careers or client engagements.', FONT_BODY)


def build_team_members(builder):
    qualified = [
        ('CA Ruchi Singh', 'ACA, B.Com', '3 Years'),
        ('CA Viraj Agarwal', 'ACA, B.Com', '3 Years'),
        ('CA Ajit Dev Pandey', 'ACA, B.Com', '2 Years'),
        ('Mr. P M Gupta', 'Diploma in Electrical Engineering, CAIIB (Part I)', 'Retired Banker as Chief Manager, UBI in 2013'),
        ('Mr. R. K. Pandey', 'BA, LLB', '26 Years'),
        ('CS Apoorva Singh', 'B.Com (H), ACS', '13 Years'),
        ('CS Urmi Chhaparia', 'B.Com, ACS', '8 Years'),
        ('Miss Shreya Pandey', 'B.Com, Qualified Company Secretary', 'Newly Qualified'),
    ]
    semi = [
        'Shruti Khemka', 'Ritika Khosala', 'Jasraj Singh', 'Shivam Agrwal', 'Vishesh Misra',
        'Harsh Jaiswal', 'Nehal Ahmad', 'Vaibhav Jaiswal', 'Shreya Gupta', 'Yash Agrawal', 'Kajal Parikh',
    ]
    builder.section('TEAM MEMBERS', 18, 'Hero', 'Page hero', 'View Partners | Support Team')
    builder.el('Kicker', 'Team Members', FONT_EYEBROW)
    builder.el('Title', 'Our People', FONT_PAGE_TITLE)
    builder.el(
        'Subtitle',
        'Qualified and semi-qualified professionals supporting engagement delivery.',
        FONT_BODY,
    )
    builder.section('TEAM MEMBERS', 18, 'Qualified Staff', 'Card grid', '')
    builder.el('Section title', 'Qualified Staff', FONT_TITLE)
    for name, qual, exp in qualified:
        builder.el('Member name', name, FONT_CARD)
        builder.el('Qualification', qual, FONT_BODY)
        builder.el('Experience', exp, FONT_BODY)
    builder.section('TEAM MEMBERS', 18, 'Semi-Qualified Staff', 'Card grid', '')
    builder.el('Section title', 'Semi-Qualified Staff', FONT_TITLE)
    for name in semi:
        builder.el('Member name', name, FONT_CARD)
        builder.el('Qualification', 'CA Intermediate (Pursuing CA Final)' if name not in ('Yash Agrawal', 'Kajal Parikh') else 'CS Executive (Pursuing CS Professional)', FONT_BODY)
    builder.section('TEAM MEMBERS', 18, 'CTA', 'Band', 'Schedule Consultation | Contact Us')
    builder.el('Title', 'Need the right team for your assignment?', FONT_TITLE)
    builder.el('Description', 'Tell us your requirement and we will connect you with the right people.', FONT_BODY)


def build_support_team(builder):
    builder.section('SUPPORT TEAM', 19, 'Hero', 'Page hero', 'Partners | Team Members')
    builder.el('Kicker', 'Support Team', FONT_EYEBROW)
    builder.el('Title', 'Support Strength', FONT_PAGE_TITLE)
    builder.el(
        'Subtitle',
        'Dedicated manpower across audit, tax, advisory, and execution support.',
        FONT_BODY,
    )
    for m in [
        '20 Audit Staff', '10 Finance & Consultancy', '8 Tax & Legal',
        '5 Government Schemes', '7 EDP Operators', '10 Support Staff',
    ]:
        builder.el('Metric', m, FONT_CARD)
    builder.section('SUPPORT TEAM', 19, 'CTA', 'Band', 'Contact Us | Schedule Consultation')
    builder.el('Title', 'Need a team aligned to your assignment?', FONT_TITLE)
    builder.el('Description', 'Contact us to discuss scope, staffing, and timelines.', FONT_BODY)


def build_clients(builder, data):
    builder.section('CLIENTS', 6, 'Hero', 'Page hero', 'Schedule Consultation | Contact Us')
    builder.el('Kicker', 'Clients', FONT_EYEBROW)
    builder.el('Title', 'Organizations We Work With', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'Corporates, MSMEs, startups, and institutions across India.', FONT_BODY)

    builder.section('CLIENTS', 6, 'Client Network', 'Logo grid', '')
    builder.el('Title', 'Client Network', FONT_TITLE)
    for logo in data['clients']:
        builder.el('Client logo', logo['name'], FONT_BODY)

    builder.section('CLIENTS', 6, 'Who We Support', 'Segments', '')
    for title in [
        'Corporates & Listed Companies',
        'MSMEs & Growing Businesses',
        'Startups',
        'Individuals & HNIs',
    ]:
        builder.el('Segment title', title, FONT_CARD)

    builder.section('CLIENTS', 6, 'CTA', 'Band', 'Schedule Consultation | Contact Us')
    builder.el('Title', 'Ready to work with us?', FONT_TITLE)
    builder.el('Description', 'Tell us about your requirements. We will outline scope and next steps.', FONT_BODY)


def build_insights(builder, data):
    builder.section('INSIGHTS', 7, 'Hero', 'Page hero', 'Schedule Consultation | Contact Us')
    builder.el('Kicker', 'Insights', FONT_EYEBROW)
    builder.el('Title', 'Articles & Updates', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'Practical notes on GST, tax, company law, audit, and compliance.', FONT_BODY)
    for cat in ['All', 'Tax', 'GST', 'Compliance', 'Audit', 'Startups', 'Advisory', 'Government Schemes']:
        builder.el('Category filter', cat, FONT_BODY)
    builder.el('Empty state', 'No articles in this category yet.', FONT_BODY)
    for art in data['insights']:
        builder.section('INSIGHTS', 7, art['title'], 'Article card', 'Read more →')
        builder.el('Category', art['category'], FONT_BODY)
        builder.el('Date', art.get('dateDisplay', ''), FONT_BODY)
        builder.el('Summary', art['summary'], FONT_BODY)
        for para in art.get('body', []):
            builder.el('Body paragraph', para, FONT_BODY)
    builder.section('INSIGHTS', 7, 'CTA', 'Band', 'Contact Us | Schedule Consultation')
    builder.el('Title', 'Need advice on a specific topic?', FONT_TITLE)
    builder.el('Description', 'Share your question and we will point you to the right next step.', FONT_BODY)


def build_other_pages(builder):
    builder.section('CAREERS', 8, 'Main', 'Page hero + text', 'Contact Us')
    builder.el('Kicker', 'Work With Us', FONT_EYEBROW)
    builder.el('Title', 'Careers', FONT_PAGE_TITLE)
    builder.el('Subtitle', 'Join our team of professionals. Explore opportunities at Dwivedi Gupta & Co.', FONT_BODY)
    builder.el(
        'Body',
        'We are always looking for talented chartered accountants and finance professionals. For current openings, please write to us at shikhar.dwivedi@dgc.ind.in or check back later.',
        FONT_BODY,
    )

    builder.section('SCHEDULE CONSULTATION', 9, 'Hero', 'sched-hero', 'Explore Services | Talk to Our Team')
    builder.el('Kicker', 'Schedule Consultation', FONT_EYEBROW)
    builder.el('Title', 'Book Your Consultation Session', FONT_PAGE_TITLE)
    builder.el(
        'Description',
        'Reserve a focused 30-minute advisory call with our CA team. Share your preferred slot and service area, and we will confirm the engagement mode within 1-2 business days.',
        FONT_BODY,
    )
    for t in ['Free 30-minute discussion', 'No obligation advisory', 'Confidential interaction']:
        builder.el('Trust chip', t, FONT_BODY)
    builder.section('SCHEDULE CONSULTATION', 9, 'Form section', 'Form', 'Request Consultation')
    builder.el('Kicker', 'Consultation Desk', FONT_EYEBROW)
    builder.el('Title', 'Share your details and preferred slot', FONT_TITLE)
    builder.el(
        'Intro',
        'Our team reviews each request personally and aligns the consultation format to your requirement.',
        FONT_BODY,
    )
    for title, desc in [
        ('Free 30-minute discussion', 'Explain your situation at no charge. We will outline how we can help before you engage us.'),
        ('Your choice of format', 'Meet at our office, speak on the phone, or join by video call.'),
        ('Experienced CAs', 'You speak with chartered accountants who handle tax, audit, and compliance work regularly.'),
    ]:
        builder.el('Why book title', title, FONT_CARD)
        builder.el('Why book text', desc, FONT_BODY)
    builder.el('Form heading', 'Request your slot', FONT_TITLE)
    builder.el('Form subheading', 'Fill in your details and preferred time. We’ll confirm via email or phone.', FONT_BODY)
    for label in ['Full name', 'Email', 'Phone', 'Service of interest', 'Preferred date', 'Preferred time', 'Additional details']:
        builder.el('Form label', label, FONT_BODY)
    for opt in [
        'Tax Planning & Compliance', 'Audit & Assurance', 'Accounting & Bookkeeping',
        'GST Advisory', 'Company Formation & Compliance', 'Financial Advisory', 'Other',
    ]:
        builder.el('Service option', opt, FONT_BODY)
    builder.el('Button', 'Request Consultation', FONT_BODY)
    builder.el('Hint', 'We’ll respond within 1–2 business days. Your information is kept confidential.', FONT_BODY)

    builder.section('CONTACT', 10, 'Hero', 'contact-hero', 'Schedule Consultation | WhatsApp')
    builder.el('Kicker', 'Contact', FONT_EYEBROW)
    builder.el('Title', 'Talk to our Chartered Accountants', FONT_PAGE_TITLE)
    for kw in ['Tax', 'Audit', 'GST', 'Company Law', 'Compliance', 'Advisory']:
        builder.el('Keyword chip', kw, FONT_BODY)
    builder.section('CONTACT', 10, 'Reach Us', 'Card', '')
    builder.el('Email', 'shikhar.dwivedi@dgc.ind.in', FONT_BODY)
    builder.el('Mobile', '+91 9721227799', FONT_BODY)
    builder.el('Landline', '0542-2502525', FONT_BODY)
    builder.el('WhatsApp', '+91 9721227799', FONT_BODY)
    builder.el('Note', 'Response within 24–48 hours on working days.', FONT_BODY)
    builder.section('CONTACT', 10, 'Send a Message', 'Form', 'Send Message')
    builder.el('Title', 'Send a Message', FONT_TITLE)
    for label in ['Name', 'Email', 'Phone', 'Subject', 'Message']:
        builder.el('Form label', label, FONT_BODY)
    for subj in ['General Enquiry', 'Tax', 'Audit', 'GST', 'Company Formation', 'Compliance', 'Other']:
        builder.el('Subject option', subj, FONT_BODY)
    builder.el('Button', 'Send Message', FONT_BODY)
    builder.el('Success message', 'Thank you. We will get back to you within 24–48 hours.', FONT_BODY)
    offices = [
        ('Varanasi (Head Office)', 'S-8/108-B-3-A Prashantpuri, M.A Road, Varanasi – 221002'),
        ('Delhi (Branch)', '62, Shrestha Vihar, Vikas Marg Extension, Delhi – 110092'),
        ('Kolkata (Branch)', 'Brijdham Housing Complex, 637 Dakshin Dari Road, Kolkata'),
        ('Bokaro (Branch)', 'C-1, 21A, 2nd Floor, City Centre, Sector-4, Bokaro Steel City'),
    ]
    builder.section('CONTACT', 10, 'Our Offices', 'Office grid', '')
    builder.el('Title', 'Our Offices', FONT_TITLE)
    for label, addr in offices:
        builder.el('Office label', label, FONT_CARD)
        builder.el('Address', addr, FONT_BODY)
        builder.el('Link', 'Directions →', FONT_BODY)

    builder.section('COMPLIANCE', 11, 'Hero', 'Page hero', '')
    builder.el('Kicker', 'Regulatory Profile', FONT_EYEBROW)
    builder.el('Title', 'Compliance Information', FONT_PAGE_TITLE)
    builder.el(
        'Subtitle',
        'Core statutory and registration details of Dwivedi Gupta & Co. for reference and verification support.',
        FONT_BODY,
    )
    builder.el('Section title', 'Firm Credentials & Statutory Details', FONT_TITLE)
    for label, val in [
        ('Year of Establishment', '2003'),
        ('Firm Registration No.', '012584C'),
        ('Category of Firm (ICAI)', 'I (One)'),
        ('Official Email', 'shikhar.dwivedi@dgc.ind.in'),
        ('Head Office', 'Varanasi'),
        ('Branch Offices', 'Kolkata, Delhi & Bokaro'),
    ]:
        builder.el(label, val, FONT_BODY)
    builder.el('Note', 'Last updated: 2026. Information is subject to updates as per regulatory changes.', FONT_BODY)

    builder.section('PRIVACY POLICY', 12, 'Hero', 'Page hero', '')
    builder.el('Kicker', 'Data Protection', FONT_EYEBROW)
    builder.el('Title', 'Privacy Policy', FONT_PAGE_TITLE)
    builder.el(
        'Subtitle',
        'How Dwivedi Gupta & Co. collects, uses, stores, and safeguards information shared through our website and services.',
        FONT_BODY,
    )
    for title in ['Information We Collect', 'How We Use Your Information', 'Data Protection & Security', 'Your Rights']:
        builder.el('Section title', title, FONT_CARD)

    builder.section('TERMS', 13, 'Hero', 'Page hero', '')
    builder.el('Kicker', 'Website Terms', FONT_EYEBROW)
    builder.el('Title', 'Terms & Conditions', FONT_PAGE_TITLE)
    builder.el(
        'Subtitle',
        'Terms governing use of this website, informational content, and interactions with Dwivedi Gupta & Co.',
        FONT_BODY,
    )
    for title in ['Website Usage', 'Professional Engagement', 'Intellectual Property', 'Liability & Jurisdiction']:
        builder.el('Section title', title, FONT_CARD)

    builder.section('DISCLAIMER', 14, 'Hero', 'Page hero', '')
    builder.el('Kicker', 'Legal Notice', FONT_EYEBROW)
    builder.el('Title', 'Disclaimer', FONT_PAGE_TITLE)
    builder.el(
        'Subtitle',
        'Important notice regarding website content, professional standards, confidentiality, and reliance on information.',
        FONT_BODY,
    )
    for title in [
        'ICAI Code of Ethics', 'Confidentiality', 'Independence',
        'Regulatory Compliance', 'No Professional Advice via Website', 'Reliance and Limitation',
    ]:
        builder.el('Card title', title, FONT_CARD)
    builder.el('Updated', 'Last updated: April 2026.', FONT_BODY)

    builder.section('SITEMAP', 15, 'Hero', 'Page hero', '')
    builder.el('Kicker', 'Site Navigation', FONT_EYEBROW)
    builder.el('Title', 'Sitemap', FONT_PAGE_TITLE)
    builder.el(
        'Subtitle',
        'Explore all public pages on the Dwivedi Gupta & Co. website in one structured view.',
        FONT_BODY,
    )
    builder.el('Section title', 'Complete page map for easier navigation', FONT_TITLE)
    for g in ['Main Pages', 'Service Detail Pages', 'Insights Library']:
        builder.el('Group title', g, FONT_CARD)


def build_footer(builder):
    builder.section('FOOTER (all pages)', 17, 'Brand column', 'Footer', 'Schedule Consultation | Contact Us')
    builder.el('Company name', 'Dwivedi Gupta & Co.', FONT_CARD)
    builder.el('Tagline', 'Chartered Accountants since 2003', FONT_BODY)
    builder.el('Email', 'shikhar.dwivedi@dgc.ind.in', FONT_BODY)
    builder.el('Mobile', '+91 9721227799', FONT_BODY)
    builder.el('Offices', 'Varanasi · Delhi · Kolkata · Bokaro', FONT_BODY)
    for link in ['About', 'Services', 'Team', 'Insights', 'Contact']:
        builder.el('Nav link', link, FONT_BODY)
    for link in ['Compliance', 'Privacy', 'Terms', 'Disclaimer']:
        builder.el('Legal link', link, FONT_BODY)
    builder.el('Social', 'LinkedIn | WhatsApp | Chat on WhatsApp | Scroll to top', FONT_BODY)
    builder.el('Copyright', '© {year} Dwivedi Gupta & Co. All rights reserved.', FONT_BODY)

    builder.section('SHARED CTA BAND (defaults)', 20, 'Defaults', 'Band', 'Schedule Consultation | Contact Us')
    builder.el('Eyebrow', 'Dwivedi Gupta & Co.', FONT_EYEBROW)
    builder.el('Title', 'Discuss your requirement with our CA team', FONT_TITLE)
    builder.el(
        'Description',
        'Book a free 30-minute consultation or send us your query. We respond within 24 to 48 working hours.',
        FONT_BODY,
    )


def _style_header_rows(ws):
    headers = ['S.no', 'Page Sub-heading', 'Design Used', 'Buttons', 'Content', None, None, 'Remarks']
    subheaders = [None, None, None, None, 'Heading', 'Font name & Size', 'Text', None]
    for c, val in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=val)
    for c, val in enumerate(subheaders, 1):
        ws.cell(row=2, column=c, value=val)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    for cell in ws[1] + ws[2]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill

    widths = {'A': 8, 'B': 28, 'C': 14, 'D': 36, 'E': 18, 'F': 42, 'G': 80, 'H': 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_page_sheet(wb, sheet_title, page_rows):
    ws = wb.create_sheet(sheet_title)
    _style_header_rows(ws)

    local_sno = 0
    for r, item in enumerate(page_rows, 3):
        local_sno += 1
        values = [
            local_sno,
            item['subheading'],
            item['design'],
            item['buttons'],
            item['element'],
            item['font'],
            item['text'],
            item['remarks'],
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    return ws


def write_workbook(builder, data):
    wb = openpyxl.Workbook()

    # Index sheet first
    index = wb.active
    index.title = '00 Index'
    index_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    index_headers = ['Sheet', 'Page Name', 'Page No', 'Content Rows']
    for c, val in enumerate(index_headers, 1):
        cell = index.cell(row=1, column=c, value=val)
        cell.font = Font(bold=True)
        cell.fill = index_header_fill
    for col, width in {'A': 28, 'B': 36, 'C': 10, 'D': 14}.items():
        index.column_dimensions[col].width = width

    # Contact Data
    cds = wb.create_sheet('Contact Data')
    contact_rows = [
        ['Section', 'Field', 'Value', 'Notes', 'Last Updated'],
        ['Contact', 'Primary Email', 'shikhar.dwivedi@dgc.ind.in', 'Site-wide', str(date.today())],
        ['Contact', 'Mobile', '+91 9721227799', 'Contact page & footer', str(date.today())],
        ['Contact', 'Landline', '0542-2502525', 'Contact page & footer', str(date.today())],
        ['Contact', 'WhatsApp', '+91 9721227799', 'wa.me/919721227799', str(date.today())],
        ['Compliance', 'Firm Registration No.', '012584C', 'Compliance page', str(date.today())],
        ['Compliance', 'Official Email', 'shikhar.dwivedi@dgc.ind.in', 'Compliance page', str(date.today())],
    ]
    for r, row in enumerate(contact_rows, 1):
        for c, val in enumerate(row, 1):
            cds.cell(row=r, column=c, value=val)
    for cell in cds[1]:
        cell.font = Font(bold=True)
        cell.fill = index_header_fill

    # Group rows by page
    by_page = {}
    for item in builder.rows:
        by_page.setdefault(item['page_name'], []).append(item)

    index_row = 2
    used_sheet_names = set()
    for page_name in builder.page_order:
        page_rows = by_page.get(page_name, [])
        base_name = sheet_name_for(page_name)
        sheet_title = base_name
        # Ensure unique sheet names within Excel 31-char limit
        if sheet_title in used_sheet_names:
            suffix = 2
            while f'{base_name[:28]}_{suffix}' in used_sheet_names:
                suffix += 1
            sheet_title = f'{base_name[:28]}_{suffix}'
        used_sheet_names.add(sheet_title)

        _write_page_sheet(wb, sheet_title, page_rows)

        page_no = page_rows[0]['page_no'] if page_rows else ''
        index.cell(row=index_row, column=1, value=sheet_title)
        index.cell(row=index_row, column=2, value=page_name)
        index.cell(row=index_row, column=3, value=page_no)
        index.cell(row=index_row, column=4, value=len(page_rows))
        index_row += 1

    meta = wb.create_sheet('Audit Meta')
    meta.append(['Generated', str(date.today())])
    meta.append(['File', 'WEBSITE_CONTENT_ON_SCREEN.xlsx / _BY_PAGE.xlsx'])
    meta.append(['Layout', 'One Excel sheet per website page'])
    meta.append(['Source', 'client/src frontend — current user-visible text'])
    meta.append(['Includes', 'Services mega-menu (3 columns, up to 4 key areas each)'])
    meta.append(['Total content rows', len(builder.rows)])
    meta.append(['Page sheets', len(builder.page_order)])
    meta.append(['Services', len(data['services'])])
    meta.append(['Insights', len(data['insights'])])
    meta.append(['Clients', len(data['clients'])])
    meta.append(['Note', 'Does not replace WEBSITE_CONTENT_AUDIT.xlsx'])

    return wb


def main():
    data = load_data()
    builder = AuditBuilder()
    build_header(builder, data)
    build_home(builder)
    build_about(builder)
    build_services(builder, data)
    build_industries(builder)
    build_team(builder)
    build_clients(builder, data)
    build_insights(builder, data)
    build_other_pages(builder)
    build_service_details(builder, data)
    build_team_members(builder)
    build_support_team(builder)
    build_footer(builder)

    wb = write_workbook(builder, data)
    saved = []
    for path in (OUT_PATH, OUT_PATH_FALLBACK):
        try:
            wb.save(path)
            saved.append(path)
        except PermissionError:
            print(f'NOTE: could not write {path.name} (file open/locked).')
    if not saved:
        alt = ROOT / f'WEBSITE_CONTENT_ON_SCREEN_{date.today().isoformat()}.xlsx'
        wb.save(alt)
        saved.append(alt)
        print(f'NOTE: saved fallback copy as {alt.name}')
    print(f'Saved {", ".join(p.name for p in saved)} ({len(builder.rows)} content rows, {len(builder.page_order)} page sheets)')


if __name__ == '__main__':
    main()
