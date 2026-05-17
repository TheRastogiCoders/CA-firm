import openpyxl
from datetime import date
from openpyxl.styles import Font, PatternFill

xlsx_path = r'c:\Users\vasur\Desktop\Projects\CA-firm\WEBSITE_CONTENT_AUDIT.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb['Content Audit']

old_snapshots = {}
updates = {
    317: 'We are always looking for talented chartered accountants and finance professionals. Email shikhar.dwivedi@dgc.ind.in',
    328: 'shikhar.dwivedi@dgc.ind.in',
    329: '+91 9721227799 (Mobile) | 0542-2502525 (Landline)',
    330: '+91 9721227799 — Chat with us (wa.me/919721227799)',
    407: 'LinkedIn (company page) | mailto:shikhar.dwivedi@dgc.ind.in | tel:+919721227799 | tel:+915422502525 (landline) | WhatsApp float (+91 9721227799)',
}

for row_num, new_text in updates.items():
    old_snapshots[row_num] = ws.cell(row=row_num, column=9).value
    ws.cell(row=row_num, column=9).value = new_text

if 'Contact Data' in wb.sheetnames:
    del wb['Contact Data']
cds = wb.create_sheet('Contact Data', 0)

csv_rows = [
    ['Section', 'Field', 'Value', 'Notes', 'Last Updated'],
    ['Contact', 'Primary Email', 'shikhar.dwivedi@dgc.ind.in', 'Shown site-wide (header ticker footer contact page)', str(date.today())],
    ['Contact', 'Mobile', '+91 9721227799', 'Shown on Contact page and footer', str(date.today())],
    ['Contact', 'Landline', '0542-2502525', 'Shown on Contact page and footer', str(date.today())],
    ['Contact', 'WhatsApp', '+91 9721227799', 'WhatsApp link: wa.me/919721227799', str(date.today())],
    ['Careers', 'Application Email', 'shikhar.dwivedi@dgc.ind.in', 'Careers page', str(date.today())],
    ['Privacy', 'Privacy Queries Email', 'shikhar.dwivedi@dgc.ind.in', 'Privacy Policy contact (when published)', str(date.today())],
    ['Compliance', 'Official Email', 'aditi.kapoor@dgc.ind.in', 'Published credentials on Compliance page', ''],
    ['Compliance', 'Year of Establishment', '2003', '', ''],
    ['Compliance', 'Firm Registration No.', '012584C', '', ''],
    ['Compliance', 'RBI Unique Code', '000293', 'Leading zeros as on Compliance page', ''],
    ['Compliance', 'Peer Review Certificate No.', '014832', 'Leading zeros as on Compliance page', ''],
    ['Compliance', 'CAG Empanelment No.', 'CR3209', '', ''],
    ['Compliance', 'ICAI Firm Category', 'I (One)', '', ''],
    ['Compliance', 'PAN', 'AAEFD6000D', '', ''],
    ['Compliance', 'GSTN', '09AAEFD6000D1Z1', '', ''],
    ['Compliance', 'Head Office', 'Varanasi', '', ''],
    ['Compliance', 'Branch Offices', 'Kolkata, Delhi, Bokaro', '', ''],
]

for r, row in enumerate(csv_rows, 1):
    for c, val in enumerate(row, 1):
        cds.cell(row=r, column=c).value = val

if 'Contact Change Log' in wb.sheetnames:
    del wb['Contact Change Log']
clog = wb.create_sheet('Contact Change Log', 1)
clog.append(['Field / Location', 'Old Value', 'New Value', 'Date'])
clog_rows = [
    ('Careers — Row 317 Body', old_snapshots[317], updates[317], str(date.today())),
    ('Contact — Row 328 Email', old_snapshots[328], updates[328], str(date.today())),
    ('Contact — Row 329 Phone', old_snapshots[329], updates[329], str(date.today())),
    ('Contact — Row 330 WhatsApp', old_snapshots[330], updates[330], str(date.today())),
    ('Footer — Row 407 Social', old_snapshots[407], updates[407], str(date.today())),
    ('CSV — Primary Email', 'vivek.gupta@dgc.ind.in', 'shikhar.dwivedi@dgc.ind.in', str(date.today())),
    ('CSV — Secondary Email', 'shikhar.dwivedi@dgc.ind.in', '(removed — single primary email)', str(date.today())),
    ('CSV — Primary Phone', '+91 94158 05906', '(removed)', str(date.today())),
    ('CSV — Secondary Phone', '+91 9721227799', 'Mobile: +91 9721227799', str(date.today())),
    ('CSV — Landline', '(not listed)', '0542-2502525', str(date.today())),
    ('CSV — WhatsApp', '(not listed)', '+91 9721227799', str(date.today())),
    ('CSV — Careers Email', 'vivek.gupta@dgc.ind.in', 'shikhar.dwivedi@dgc.ind.in', str(date.today())),
    ('CSV — Privacy Email', 'vivek.gupta@dgc.ind.in', 'shikhar.dwivedi@dgc.ind.in', str(date.today())),
    ('CSV — RBI Unique Code', '293', '000293', str(date.today())),
    ('CSV — Peer Review Cert', '14832', '014832', str(date.today())),
    ('Website Header ticker', 'connect@dgcindia.com / info@dgcindia.com / +91 94500 00000', 'shikhar.dwivedi@dgc.ind.in / +91 9721227799 / 0542-2502525', str(date.today())),
]
for row in clog_rows:
    clog.append(list(row))

header_font = Font(bold=True)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
for sheet in [cds, clog]:
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

wb.save(xlsx_path)
print('Excel updated successfully.')
for row_num in sorted(old_snapshots):
    print(f'Row {row_num}: OLD -> {old_snapshots[row_num]}')
    print(f'         NEW -> {updates[row_num]}')
