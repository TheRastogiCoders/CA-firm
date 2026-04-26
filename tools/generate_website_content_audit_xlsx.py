#!/usr/bin/env python3
"""
Build WEBSITE_CONTENT_AUDIT.xlsx — content inventory table matching the audit layout
(merged headers, Page Name + Page No vertical merges per page).
Typography column reflects the live site (DM Sans from index.html / index.css).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "WEBSITE_CONTENT_AUDIT.xlsx"

# --- Font tokens (match client/src/index.css & home-sections.css) ---
F_LABEL = "DM Sans, 600, ~0.78rem (.hero-label)"
F_TAG = "DM Sans, 500–600, clamp(0.92–1rem) (.hero-tagline)"
F_H1_HERO = "DM Sans, 600, clamp(1.95–3.25rem) (.hero-title)"
F_H1_HERO_SM = "DM Sans, 600, smaller hero (.hero-title-sm)"
F_HERO_DESC = "DM Sans, 400, clamp(0.98–1.08rem) (.hero-desc)"
F_EYEBROW = "DM Sans, 600, ~0.8125rem (section eyebrow)"
F_H2_HOME = "DM Sans, 600, clamp(1.5–2.25rem) (.about-title / .services-title)"
F_H2_PRESENCE = "DM Sans, 600, clamp(1.75–2.25rem) (.presence-title)"
F_H3_CARD = "DM Sans, 600, ~1.0625–1.25rem (cards)"
F_BODY = "DM Sans, 400, ~1rem (body)"
F_STAT = "DM Sans, 600, display stat values"
F_CTA_BAND = "DM Sans, 600 (.home-hero-2-title / .btn)"
F_BTN = "DM Sans, 600, 1rem (.btn / hero buttons)"
F_PAGE_H1 = "DM Sans, 600, clamp(1.9–3rem) (.page-title)"
F_PAGE_SUB = "DM Sans, 400, clamp(1–1.12rem) (.page-subtitle)"
F_SECTION_H2 = "DM Sans, 600, clamp(1.75–2.5rem) (.section-title)"
F_SCHED = "DM Sans, 600 / 400 (.sched-hero-*, form)"
F_COMPLIANCE = "DM Sans, 600 / 400 (.compliance-*)"

THIN = Side(style="thin", color="000000")


def add_row(
    rows: list,
    page_name: str,
    page_no: int,
    sub: str,
    design: str,
    buttons: str,
    heading: str,
    font_spec: str,
    text: str,
    remarks: str = "",
):
    rows.append(
        {
            "page_name": page_name,
            "page_no": page_no,
            "sub": sub,
            "design": design,
            "buttons": buttons,
            "heading": heading,
            "font": font_spec,
            "text": text,
            "remarks": remarks,
        }
    )


SERVICE_AUDIENCE = {
    "tax-compliance": "Ideal for companies, firms, and professionals managing periodic tax obligations.",
    "audit-assurance": "Suitable for entities requiring statutory confidence, controls review, and reporting clarity.",
    "gst-advisory": "Best for businesses with ongoing GST filings, reconciliations, and notices.",
    "corporate-law": "Useful for companies and LLPs needing ongoing MCA/ROC and governance support.",
    "company-formation": "Designed for founders launching a new business entity and setting compliance foundations.",
    "financial-consulting": "Recommended for businesses evaluating investments, restructuring, or transaction decisions.",
    "project-finance": "Helpful for promoters and enterprises seeking debt support for expansion projects.",
    "government-schemes-advisory": "For businesses exploring eligible central/state subsidies and incentive programs.",
}


def build_services_data():
    """Mirrors client/src/data/servicesData.js (frontend source of truth)."""
    return [
        {
            "slug": "tax-compliance",
            "title": "Tax & Compliance",
            "short": "Strategic tax planning, direct and indirect tax compliance, and representation before authorities to optimise your tax position while meeting all regulatory requirements.",
            "long": "Our Tax & Compliance practice helps businesses and individuals navigate the complexities of Indian tax laws. We provide strategic tax planning to minimise liability within the legal framework, ensure timely compliance with income tax, TDS, and indirect tax obligations, and represent you before tax authorities when required. Our team stays abreast of amendments and judicial developments to give you accurate, practical advice.",
            "areas": "Income tax planning & returns; TDS & TCS compliance; Indirect taxes (GST coordination); Tax representation & appeals; International tax aspects",
        },
        {
            "slug": "audit-assurance",
            "title": "Audit & Assurance",
            "short": "Statutory audits, internal audits, tax audits, and assurance services that strengthen financial controls, transparency, and stakeholder confidence.",
            "long": "We deliver a full range of audit and assurance services tailored to your size, sector, and regulatory requirements. Our statutory audits comply with applicable standards and provide assurance to shareholders and regulators. We also conduct internal audits, tax audits, and stock audits to strengthen controls, identify risks, and improve processes. Our reports are clear, actionable, and aligned with best practices.",
            "areas": "Statutory audit; Internal audit; Tax audit; Stock audit; Management audit; Assurance & agreed-upon procedures",
        },
        {
            "slug": "gst-advisory",
            "title": "GST Advisory",
            "short": "End-to-end GST support including registration, returns, refunds, and litigation. We help you stay compliant and leverage GST benefits effectively.",
            "long": "GST impacts every business. We offer end-to-end GST advisory—from registration and classification to monthly/quarterly returns, refunds, and litigation support. We help you optimise input tax credit, maintain robust documentation, and respond to notices and assessments. Our team keeps you updated on rate changes, compliance deadlines, and opportunities under the GST regime.",
            "areas": "GST registration & amendments; Returns (GSTR-1, GSTR-3B, annual); Refunds & reconciliation; Litigation & dispute resolution; Compliance reviews & health checks",
        },
        {
            "slug": "corporate-law",
            "title": "Corporate Law",
            "short": "Company law advisory, corporate governance, secretarial services, and regulatory filings to keep your organisation compliant and well-governed.",
            "long": "We advise on company law, corporate governance, and secretarial compliance. Our services include board and general meeting support, maintenance of statutory registers, filing of forms and returns with the Registrar of Companies, and guidance on SEBI and other regulations where applicable. We help listed and unlisted companies, LLPs, and other entities stay compliant and governance-ready.",
            "areas": "Company law & governance; Secretarial services; ROC & MCA filings; Board & general meetings; Compliance calendars",
        },
        {
            "slug": "company-formation",
            "title": "Company Formation",
            "short": "Incorporation of companies and LLPs, ROC filings, and ongoing corporate compliance so you can focus on running your business.",
            "long": "Starting a company or LLP involves multiple steps and ongoing compliance. We handle incorporation, drafting of MOA/AOA or LLP agreement, name approval, and registration. Post-incorporation, we support you with annual filings, director KYC, changes in directorship or capital, and other ROC-related compliances so your entity remains in good standing and you can focus on business.",
            "areas": "Company & LLP incorporation; ROC filings & amendments; Annual returns & financial statements; Director KYC & changes; Strike-off & restoration support",
        },
        {
            "slug": "financial-consulting",
            "title": "Financial Consulting",
            "short": "Business valuation, due diligence, restructuring, and financial advisory to support mergers, investments, and strategic decisions.",
            "long": "Our Financial Consulting practice supports strategic decisions with rigorous analysis and independent advice. We provide business and asset valuation, financial and commercial due diligence for M&A and investments, and restructuring advisory. Whether you are raising capital, acquiring a business, or reorganising your group, we deliver clarity and credibility to the numbers.",
            "areas": "Business & asset valuation; Financial due diligence; M&A advisory; Restructuring & demerger; Fundraising support",
        },
        {
            "slug": "project-finance",
            "title": "Project Finance",
            "short": "Financial structuring, debt syndication, and banking support for projects. We help you secure funding and structure deals that work.",
            "long": "We support projects and infrastructure with financial structuring, debt syndication, and banking liaison. Our team helps prepare bankable financial models, information memorandums, and documentation required by lenders. We work with promoters, banks, and institutions to align structure, covenants, and timelines so projects can achieve financial closure and stay on track.",
            "areas": "Financial structuring; Debt syndication; Banking support & documentation; Project financial models; Lender due diligence support",
        },
        {
            "slug": "government-schemes-advisory",
            "title": "Government Schemes Advisory",
            "short": "Awareness, planning, and implementation support for Central and State Government schemes across sectors to help you access benefits and incentives.",
            "long": "Central and State Governments run numerous schemes for industry, export, R&D, and sector-specific incentives. We help you understand eligibility, documentation, and compliance requirements for relevant schemes. Our advisory includes scheme selection, application support, and ongoing compliance so you can access subsidies, interest subventions, and other benefits in a timely, compliant manner.",
            "areas": "Scheme awareness & selection; Application & documentation support; Compliance for incentives; State & Central scheme mapping",
        },
    ]


def build_insights():
    """Mirrors client/src/data/insightsData.js (summaries + body paragraphs)."""
    return [
        {
            "title": "GST Annual Return: Checklist for FY 2024-25",
            "cat": "GST",
            "date": "15 Feb 2025",
            "summary": "Key points to keep in mind while filing your GST annual return and reconciliation.",
            "body": (
                "The GST annual return (GSTR-9) and annual reconciliation (GSTR-9C for turnover above the threshold) are critical compliances for every registered person. Before you file, ensure your books are reconciled with the returns already filed during the year—GSTR-1, GSTR-3B, and the e-invoice/ e-way bill data where applicable.\n\n"
                "Reconcile outward supplies with GSTR-1 and GSTR-3B. Match input tax credit (ITC) with GSTR-2B and your purchase register. Discrepancies should be identified and either corrected through amendments or explained in the annual return. Keep supporting documents ready for any audit or notice.\n\n"
                "Due dates are typically extended by the GST Council; check the latest notifications. Late fees and interest apply for delayed filing. If you need help with reconciliation or representation, reach out to your CA or our team for support."
            ),
        },
        {
            "title": "Tax Saving Options for Salaried Individuals",
            "cat": "Tax",
            "date": "20 Jan 2025",
            "summary": "Section 80C, 80D, HRA, and other deductions you can use before March 31.",
            "body": (
                "Salaried individuals can reduce their tax outgo by using the right deductions and exemptions. Section 80C allows a deduction of up to ₹1.5 lakh for investments in PPF, ELSS, EPF, life insurance premium, NSC, tuition fees, and repayment of housing loan principal. Plan your investments before 31 March to claim the benefit for the current financial year.\n\n"
                "Section 80D offers deduction for health insurance premium (self, family, parents)—up to ₹25,000 or ₹50,000 for senior citizens. HRA exemption depends on salary structure, rent paid, and city of residence. Leave Travel Allowance (LTA), interest on housing loan (Section 24), and the new tax regime options should also be evaluated.\n\n"
                "We recommend a quick review with your CA to ensure you are not missing any claim and to plan investments and documents (rent receipts, insurance premium paid, etc.) well before the deadline."
            ),
        },
        {
            "title": "MCA Compliance Calendar: Key Due Dates",
            "cat": "Compliance",
            "date": "10 Jan 2025",
            "summary": "ROC filings, AOC-4, MGT-7, and other company law compliances at a glance.",
            "body": (
                "Companies and LLPs must adhere to several MCA/ROC due dates through the year. Key annual compliances include filing of financial statements (AOC-4 for companies) and annual return (MGT-7), typically within 30 days and 60 days respectively from the conclusion of the AGM. For LLPs, Form 8 (Statement of Account) and Form 11 (Annual Return) have their own deadlines.\n\n"
                "Event-based filings include changes in directors, registered office, capital, and charges. ADT-1 (appointment of auditor), MGT-14 (for certain resolutions), and DIR-3 KYC for directors are also part of the compliance calendar. Missing dates can lead to additional fees and disqualification risks.\n\n"
                "Maintain an internal compliance calendar or work with your secretarial team/CA to ensure nothing is missed. We help clients with ROC filings and compliance tracking across Varanasi, Delhi, Kolkata, and Bokaro."
            ),
        },
        {
            "title": "Startup India: Benefits and Registration",
            "cat": "Startups",
            "date": "5 Dec 2024",
            "summary": "Eligibility, tax benefits, and how to register under the Startup India scheme.",
            "body": (
                "Startup India aims to support innovation and growth through tax benefits, easier compliance, and access to schemes. Eligibility criteria include incorporation as a private limited company, LLP, or partnership; turnover not exceeding ₹100 crore; and recognition as an eligible startup (typically within 10 years from incorporation).\n\n"
                "Benefits include income tax exemption for three consecutive years in a block of seven years (subject to conditions), relaxation in compliance under the Companies Act, and fast-tracking of patent applications. Registration is done through the Startup India portal with supporting documents and a recommendation/ recognition certificate.\n\n"
                "We assist startups with incorporation, DPIIT recognition, and ongoing compliance so they can focus on building their business while staying eligible for benefits."
            ),
        },
        {
            "title": "TDS Compliance for Businesses: What to Track",
            "cat": "Tax",
            "date": "28 Nov 2024",
            "summary": "Deduction, deposit, and return filing—a quick guide to staying TDS-compliant.",
            "body": (
                "Tax Deducted at Source (TDS) applies to a wide range of payments—salary, rent, professional fees, contract payments, interest, and more. As a deductor, you must deduct tax at the prescribed rates, deposit it within the due dates (usually 7th of the following month), and file TDS returns (e.g. Q1–Q4) in Form 24Q, 26Q, 27Q, etc.\n\n"
                "Late deduction or deposit attracts interest and penalties. Issuing TDS certificates (Form 16, 16A) to deductees on time is also mandatory. Use the TRACES portal for statements, corrections, and compliance status.\n\n"
                "We help businesses set up TDS processes, reconcile TDS data with books, and file returns so that both deductors and deductees remain compliant."
            ),
        },
        {
            "title": "Internal Audit: Best Practices for Mid-Size Firms",
            "cat": "Audit",
            "date": "15 Nov 2024",
            "summary": "How to design and run an internal audit function that adds value.",
            "body": (
                "Internal audit is not just about checking boxes—it should strengthen controls, identify risks, and improve processes. For mid-size companies, a risk-based audit plan aligned with business objectives works best. Focus on high-impact areas: revenue cycle, procurement, payroll, IT controls, and regulatory compliance.\n\n"
                "Documentation, independence, and clear reporting to the audit committee or management are essential. Use sampling and data analytics where possible to improve coverage and insight. Follow-up on prior findings to ensure closure.\n\n"
                "We offer internal audit services tailored to your size and sector, with practical recommendations and follow-through."
            ),
        },
        {
            "title": "GST Refunds: How to Claim and What to Expect",
            "cat": "GST",
            "date": "20 Oct 2024",
            "summary": "Eligibility, documentation, and timelines for export and inverted duty refunds.",
            "body": (
                "Refunds under GST can arise from exports (zero-rated supplies), inverted duty structure, or other specified situations. The process involves filing the refund application on the portal (RFD-01 or through the return), uploading supporting documents, and responding to any deficiency memos or notices.\n\n"
                "Export refunds typically require shipping bills, tax payment documents, and bank realisation. Inverted duty refunds need a statement of inward and outward supplies. Timelines and interest on delayed refunds are prescribed by law.\n\n"
                "We assist businesses with refund applications, reconciliation, and representation before authorities to expedite and secure refunds."
            ),
        },
        {
            "title": "Company Incorporation: A Practical Checklist",
            "cat": "Compliance",
            "date": "5 Oct 2024",
            "summary": "Documents, name approval, and post-incorporation steps for new companies.",
            "body": (
                "Incorporating a company involves name reservation (RUN or SPICe+), drafting MOA and AOA, obtaining DIN and DSC for directors, and filing the incorporation form. The MCA SPICe+ form combines name approval, incorporation, and several other applications (PAN, TAN, EPFO, ESIC) in one go.\n\n"
                "Post-incorporation, you need to open a bank account, get a seal (if required), hold the first board meeting, and comply with ongoing ROC and tax filings. Keep identity and address proofs of directors and subscribers ready, and ensure the registered office address is valid.\n\n"
                "We handle end-to-end incorporation and post-incorporation compliance so you can start operations without missing regulatory steps."
            ),
        },
        {
            "title": "Due Diligence for M&A: What Buyers and Sellers Should Know",
            "cat": "Advisory",
            "date": "18 Sep 2024",
            "summary": "Financial, tax, and legal checks that matter in mergers and acquisitions.",
            "body": (
                "Due diligence helps buyers assess risks and value, and helps sellers prepare for a smooth transaction. Financial due diligence covers quality of earnings, working capital, debt, and contingent liabilities. Tax due diligence looks at past returns, pending assessments, and potential exposure.\n\n"
                "Legal and compliance diligence covers contracts, litigation, IP, and regulatory approvals. The scope can be tailored to the deal size and risk appetite. A well-run process reduces surprises and supports negotiation and documentation.\n\n"
                "We provide financial and tax due diligence for acquisitions, investments, and divestments, with clear reports and practical recommendations."
            ),
        },
        {
            "title": "Government Schemes for MSMEs: A Quick Overview",
            "cat": "Government Schemes",
            "date": "1 Sep 2024",
            "summary": "Central and state schemes that MSMEs can tap for support and incentives.",
            "body": (
                "MSMEs can access a variety of Central and State schemes—subsidies on interest, capital investment, quality certification, and export promotion, to name a few. Udyam registration is the first step for many of these; eligibility and documentation vary by scheme.\n\n"
                "Staying updated on new schemes and amendments is important. Application processes are often online; missing deadlines can mean losing out on benefits. We help MSMEs identify relevant schemes, prepare applications, and maintain compliance for continued eligibility.\n\n"
                "Reach out to us for a discussion on which schemes fit your business and how we can support you through the process."
            ),
        },
    ]


def build_all_rows() -> list:
    rows: list = []
    corporate_values = [
        (
            "Partnership",
            "Instead of being a distant service provider, we collaborate with our clients in all our engagements, work with them as a team and take ownership and responsibility of things, to create long lasting partnerships.",
        ),
        (
            "Integrity",
            "Our services are aimed at protecting our client's interests. By adopting transparent processes and adhering to highest ethical standards, we ensure client confidentiality and our own credibility. Whilst collaborating with our clients, we remain absolutely independent to deliver unbiased opinions.",
        ),
        (
            "Passion",
            "We are passionate for our client's success. By creating a highly stimulating work environment, working with utmost dedication and commitment and focusing on delivery and execution, we perform to not just satisfy but delight our clients.",
        ),
        (
            "Excellence",
            "By continually focusing on quality and deploying best practices, we bring excellence in our work, add value for our clients and strive to enter the realm of supremacy.",
        ),
    ]

    # ----- HOME (Page 1) -----
    p, n = "HOME", 1
    for sn, slide in [
        (
            1,
            (
                "DWIVEDI GUPTA & CO.",
                "Chartered Accountants | Advisors | Consultants",
                "Chartered Accountants for Tax, Audit & Advisory",
                "We help businesses with tax filing, audits, compliance, and financial advisory services.",
            ),
        ),
        (
            2,
            (
                "DWIVEDI GUPTA & CO.",
                "Assurance | Taxation | Advisory | Consulting",
                "Compliance and Growth Support for Businesses",
                "From statutory compliance to strategic advisory, we support corporates, SMEs, and institutions.",
            ),
        ),
        (
            3,
            (
                "DWIVEDI GUPTA & CO.",
                "Est. 2003 — 20+ Years of Excellence",
                "Practical Financial and Regulatory Guidance",
                "Our team delivers clear, partner-led support in tax, audit, corporate law, and finance.",
            ),
        ),
    ]:
        base = f"Hero Slide {sn}"
        btn = "Get Consultation | Contact Us" if sn == 1 else ""
        add_row(rows, p, n, base, "Hero", btn, "Label", F_LABEL, slide[0])
        add_row(rows, p, n, base, "Hero", "", "Tagline", F_TAG, slide[1])
        add_row(rows, p, n, base, "Hero", "", "Title", F_H1_HERO, slide[2])
        add_row(rows, p, n, base, "Hero", "", "Description", F_HERO_DESC, slide[3])

    add_row(rows, p, n, "About the Firm", "Section", "", "Eyebrow", F_EYEBROW, "Who We Are")
    add_row(rows, p, n, "About the Firm", "Section", "", "Title", F_H2_HOME, "About the Firm")
    add_row(
        rows,
        p,
        n,
        "About the Firm",
        "Stats row",
        "",
        "Stats",
        F_STAT,
        "2003 (Established) | 4 (Offices) | 20+ (Years Experience)",
    )
    add_row(
        rows,
        p,
        n,
        "About the Firm",
        "Section",
        "Learn More About Us",
        "Lead paragraph",
        F_BODY,
        "Founded in 2003, Dwivedi Gupta & Co. (DGC) is a Chartered Accountants firm delivering assurance, taxation, and advisory services with a partner-led approach and deep regulatory insight.",
    )
    add_row(
        rows,
        p,
        n,
        "About the Firm",
        "Section",
        "",
        "Paragraph",
        F_BODY,
        "We work as long-term partners, not distant service providers. Every engagement is tailored to your business model, growth stage, and compliance requirements rather than following a one-size-fits-all process.",
    )
    add_row(
        rows,
        p,
        n,
        "About the Firm",
        "Section",
        "",
        "Paragraph",
        F_BODY,
        "Headquartered in Varanasi with branch offices in Delhi, Kolkata, and Bokaro, our team combines local accessibility with national capability. Under the direction of 7 partners and qualified professionals, we focus on practical, timely, and high-quality outcomes.",
    )
    add_row(
        rows,
        p,
        n,
        "About the Firm",
        "Section",
        "",
        "Bullets",
        F_BODY,
        "RBI Registered and CAG Empanelled credentials | Comprehensive services: Tax, Audit, Compliance, Advisory, and Finance | Strong client relationships built on integrity and responsiveness",
    )

    add_row(rows, p, n, "Core Services", "Card grid", "View All Services", "Eyebrow", F_EYEBROW, "What We Offer")
    add_row(rows, p, n, "Core Services", "Card grid", "", "Title", F_H2_HOME, "Our Core Services")
    add_row(
        rows,
        p,
        n,
        "Core Services",
        "Card grid",
        "",
        "Intro",
        F_BODY,
        "Comprehensive assurance, tax, advisory, and consulting solutions tailored to your business needs.",
    )
    for title, desc in [
        (
            "Tax & Regulatory Services",
            "Strategic tax planning, compliance, and representation services to optimize tax efficiency while ensuring full regulatory compliance.",
        ),
        (
            "Audit & Assurance",
            "Comprehensive audit solutions including statutory audits, internal audits, tax audits, stock audits, and management audits to strengthen financial controls and transparency.",
        ),
        (
            "Corporate Law & Compliance",
            "Company incorporation, corporate governance support, secretarial services, regulatory filings, and legal compliance assistance.",
        ),
        (
            "Advisory & Consulting",
            "Business advisory services across taxation, regulatory frameworks, LLPs, trusts, and financial regulations with research-based professional insights.",
        ),
        (
            "Finance & Project Consultancy",
            "End-to-end financial consulting including project finance, debt syndication, banking support, and financial structuring for businesses.",
        ),
        (
            "Government Schemes Consultancy",
            "Awareness, planning, and implementation support for Central and State Government schemes across multiple sectors.",
        ),
    ]:
        add_row(rows, p, n, "Core Services", "Service card", "", "Card title", F_H3_CARD, title)
        add_row(rows, p, n, "Core Services", "Service card", "", "Card text", F_BODY, desc)

    add_row(rows, p, n, "Why Choose Us", "Points grid", "", "Eyebrow", F_EYEBROW, "Our Edge")
    add_row(rows, p, n, "Why Choose Us", "Points grid", "", "Title", F_H2_HOME, "Why Choose Us")
    add_row(
        rows,
        p,
        n,
        "Why Choose Us",
        "Points grid",
        "",
        "Intro",
        F_BODY,
        "Decades of experience, a client-first approach, and end-to-end CA services that drive compliance and growth.",
    )
    for title, desc in [
        ("Expert Team", "Qualified CAs and professionals with deep expertise in tax, audit, and compliance."),
        ("Full-Service", "From incorporation to annual filings, audit, and advisory—we cover it all under one roof."),
        ("Partner-Led", "Partner-led engagement ensures quality assurance and direct access to senior expertise."),
        ("Timely & Reliable", "We meet deadlines and keep you updated so you never miss a compliance date."),
        ("RBI & CAG Empanelled", "Recognized credentials that reflect our standards and capability to serve institutions."),
        ("Multi-City Presence", "Head office in Varanasi with branches in Delhi, Kolkata, and Bokaro for nationwide reach."),
    ]:
        add_row(rows, p, n, "Why Choose Us", "Point", "", title, F_H3_CARD, desc)

    add_row(rows, p, n, "Industries", "Image grid", "Explore Industries", "Eyebrow", F_EYEBROW, "Sector Expertise")
    add_row(rows, p, n, "Industries", "Image grid", "", "Title", F_H2_HOME, "Industries We Serve")
    add_row(
        rows,
        p,
        n,
        "Industries",
        "Image grid",
        "",
        "Intro",
        F_BODY,
        "We provide services across multiple industries with sector-specific knowledge and compliance support.",
    )
    for ind in [
        "Manufacturing & Infrastructure",
        "Banking & Financial Institutions",
        "Real Estate & Construction",
        "Trading & Export Businesses",
        "Government & Public Sector",
        "SMEs & Startups",
        "Non-Profit Organizations",
    ]:
        add_row(rows, p, n, "Industries", "Tile", "", "Industry", F_H3_CARD, ind)

    add_row(rows, p, n, "Our Presence", "Cards", "", "Eyebrow", F_EYEBROW, "Where We Are")
    add_row(rows, p, n, "Our Presence", "Cards", "", "Title", F_H2_PRESENCE, "Our Presence")
    add_row(
        rows,
        p,
        n,
        "Our Presence",
        "Cards",
        "",
        "Intro",
        F_BODY,
        "Headquartered in Varanasi with branch offices across key cities to serve you locally.",
    )
    add_row(
        rows,
        p,
        n,
        "Our Presence",
        "HQ card",
        "",
        "Head Office – Varanasi",
        F_H3_CARD,
        "S-8/108-B-3-A Prashantpuri, M.A Road, Varanasi – 221002. Our flagship office and central hub for assurance, tax, and advisory services.",
    )
    for city, addr, note in [
        ("Delhi", "62, Shrestha Vihar, Vikas Marg Extension, Delhi – 110092", "Reach us for audit, tax & compliance in NCR."),
        (
            "Kolkata",
            "Brijdham Housing Complex, 637 Dakshin Dari Road, 5th Floor Flat-5E, Building No 16-C, Kolkata, West Bengal",
            "Serving Eastern India with local expertise.",
        ),
        (
            "Bokaro",
            "C-1, 21A, 2nd Floor, City Centre, Sector-4, Bokaro Steel City, Jharkhand",
            "Industry-focused CA services in the region.",
        ),
    ]:
        add_row(
            rows,
            p,
            n,
            "Our Presence",
            "Branch card",
            "",
            f"Branch – {city}",
            F_BODY,
            f"{addr} — {note}",
        )
    add_row(
        rows,
        p,
        n,
        "Our Presence",
        "Cards",
        "",
        "Tagline",
        F_BODY,
        "We serve clients across India with strong regional expertise and national capabilities.",
    )

    add_row(rows, p, n, "CTA band", "Gradient CTA", "Schedule a Consultation", "Eyebrow", F_EYEBROW, "Expert guidance since 2003")
    add_row(
        rows,
        p,
        n,
        "CTA band",
        "Gradient CTA",
        "",
        "Title",
        F_CTA_BAND,
        "Looking for Reliable Financial & Advisory Experts?",
    )
    add_row(
        rows,
        p,
        n,
        "CTA band",
        "Gradient CTA",
        "",
        "Description",
        F_BODY,
        "Partner with Dwivedi Gupta & Co. for professional guidance, compliance confidence, and business growth.",
    )

    add_row(rows, p, n, "Clients preview", "Logo strip", "View Our Clients", "Eyebrow", F_EYEBROW, "Who We Work With")
    add_row(rows, p, n, "Clients preview", "Logo strip", "", "Title", F_H2_HOME, "Trusted by Leading Organizations")
    add_row(
        rows,
        p,
        n,
        "Clients preview",
        "Logo strip",
        "",
        "Intro",
        F_BODY,
        "Corporates, banks, and institutions across India rely on us for their financial and advisory needs.",
    )
    add_row(
        rows,
        p,
        n,
        "Clients preview",
        "Logo strip",
        "",
        "Caption",
        F_BODY,
        "Representing leading names across banking, manufacturing, and institutions.",
    )

    # ----- ABOUT (2) -----
    p, n = "ABOUT US", 2
    add_row(rows, p, n, "Hero", "Hero", "Schedule Consultation | Contact Us", "Label", F_LABEL, "DWIVEDI GUPTA & CO.")
    add_row(rows, p, n, "Hero", "Hero", "", "Tagline", F_TAG, "Chartered Accountants | Advisors | Consultants")
    add_row(rows, p, n, "Hero", "Hero", "", "Title", F_H1_HERO, "About Dwivedi Gupta & Co.")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Hero",
        "",
        "Description",
        F_HERO_DESC,
        "A reputed firm providing taxation, audit & assurance, advisory, and financial consulting services with over two decades of experience.",
    )
    add_row(rows, p, n, "About the Firm", "Section", "View Compliance Information", "Eyebrow", F_EYEBROW, "Who We Are")
    add_row(rows, p, n, "About the Firm", "Section", "", "Title", F_H2_HOME, "About the Firm")
    add_row(rows, p, n, "About the Firm", "Section", "", "Stats", F_STAT, "2003 | 4 Offices | 20+ Years Experience")
    add_row(
        rows,
        p,
        n,
        "About the Firm",
        "Section",
        "",
        "Lead",
        F_BODY,
        "Founded in 2003, Dwivedi Gupta & Co. (DGC) is a Chartered Accountants firm providing Assurance, Taxation and Advisory/Consulting services. Revered for our professional ethos and technical expertise, drawn on perspicacity of over two decades and a team of highly competent professionals, we provide efficacious solutions to our client's needs, running deep into their engagements, understanding need and products.",
    )
    add_row(
        rows,
        p,
        n,
        "About the Firm",
        "Section",
        "",
        "Paragraph",
        F_BODY,
        "Our philosophy is of partnering with our clients and not being a distant service provider. Since businesses are inherently different, we tailor our services to meet client's specific needs and banish the 'one-size-fits-all' standardization.",
    )
    add_row(
        rows,
        p,
        n,
        "About the Firm",
        "Section",
        "",
        "Paragraph",
        F_BODY,
        "We recruit, train, motivate and retain highly capable and sharpest talent, who bring quality in their work and deliver the best solutions. Headquartered in Varanasi with branches at Kolkata, Delhi and Bokaro, at each center we aim to deploy state-of-art infrastructure, best practices and people development programs. Under the able direction of 7 partners & other qualified CA Staff, DGC's team strength is uniquely positioned to provide you quality opinions and services. Our interdisciplinary approach renders to give you seamless value.",
    )
    add_row(
        rows,
        p,
        n,
        "About the Firm",
        "Section",
        "",
        "Paragraph",
        F_BODY,
        "Serving to the wider business community for more than two decades, we enjoy unparalleled reputation and respect of our clients, who trust and rely on us for our expertise and professionalism.",
    )
    add_row(rows, p, n, "Vision & Mission", "Cards", "", "Eyebrow", F_EYEBROW, "Our Direction")
    add_row(
        rows,
        p,
        n,
        "Vision & Mission",
        "Cards",
        "",
        "Section title",
        F_H2_HOME,
        "Vision & Mission",
    )
    add_row(
        rows,
        p,
        n,
        "Vision & Mission",
        "Cards",
        "",
        "Intro",
        F_BODY,
        "Guiding principles that define our commitment to excellence and client success.",
    )
    add_row(
        rows,
        p,
        n,
        "Vision & Mission",
        "Cards",
        "",
        "Vision",
        F_H3_CARD,
        "To become a trusted leader in professional financial and advisory services through excellence, innovation, and integrity.",
    )
    add_row(
        rows,
        p,
        n,
        "Vision & Mission",
        "Cards",
        "",
        "Mission",
        F_H3_CARD,
        "Provide high-quality services with transparency; deliver innovative financial solutions; maintain ethical standards; build lasting client relationships; and continuously enhance our expertise and capabilities.",
    )
    add_row(rows, p, n, "Core Values", "Cards", "", "Eyebrow", F_EYEBROW, "What We Stand For")
    add_row(rows, p, n, "Core Values", "Cards", "", "Title", F_H2_HOME, "Core Values")
    add_row(
        rows,
        p,
        n,
        "Core Values",
        "Cards",
        "",
        "Intro",
        F_BODY,
        "Our philosophy, principles and values are so strongly weaved in our culture fabric that our beliefs are shared amongst all and which helps us earn our client's trust and respect.",
    )
    for t, d in corporate_values:
        add_row(rows, p, n, "Core Values", "Card", "", t, F_BODY, d)
    add_row(rows, p, n, "Our People", "Section", "", "Eyebrow", F_EYEBROW, "Our People")
    add_row(rows, p, n, "Our People", "Section", "", "Title", F_H2_HOME, "Our People")
    add_row(
        rows,
        p,
        n,
        "Our People",
        "Section",
        "",
        "Intro",
        F_BODY,
        "We recruit, train, motivate and retain highly capable and sharpest talent, who bring quality in their work and deliver the best solutions. We nurture our people and turn them into our assets. The firm has adequate manpower and infrastructure to undertake specialized assignments from Government, Semi-Government, Financial Institutions, Insurance Companies, Bank and Corporate Sectors.",
    )
    add_row(rows, p, n, "Our People", "Section", "", "Stats", F_STAT, "20 Audit Staff | 10 Finance & Consultancy | 8 Tax & Legal | 5 Govt Schemes | 7 E.D.P. | 10 Semi-skilled / Peons")
    add_row(
        rows,
        p,
        n,
        "Our People",
        "Section",
        "",
        "Note",
        F_BODY,
        "Apart from qualified CA and CS staff, our firm has strategic alliance with various expert professionals who are engaged based on assignment requirements.",
    )
    add_row(rows, p, n, "Strengths", "List", "", "Eyebrow", F_EYEBROW, "Our Edge")
    add_row(rows, p, n, "Strengths", "List", "", "Title", F_H2_HOME, "Our Strengths")
    for s in [
        "20+ Years Experience",
        "Experienced Professionals",
        "Multi-City Presence",
        "Diverse Industry Expertise",
        "Strong Client Relationships",
        "Partner-Led Approach",
        "Ethical Practices",
    ]:
        add_row(rows, p, n, "Strengths", "List", "", "Strength", F_BODY, s)
    add_row(rows, p, n, "Industries", "Grid", "", "Eyebrow", F_EYEBROW, "Sector Expertise")
    add_row(rows, p, n, "Industries", "Grid", "", "Title", F_H2_HOME, "Industries We Serve")
    add_row(
        rows,
        p,
        n,
        "Industries",
        "Grid",
        "",
        "Intro",
        F_BODY,
        "We provide services across multiple industries with sector-specific knowledge and compliance support.",
    )
    for ind in [
        "Manufacturing & Infrastructure",
        "Banking & Financial Institutions",
        "Real Estate & Construction",
        "Trading & Export Businesses",
        "Government & Public Sector",
        "SMEs & Startups",
        "Non-Profit Organizations",
    ]:
        add_row(rows, p, n, "Industries", "Grid tile", "", "Industry title", F_H3_CARD, ind)
    add_row(
        rows,
        p,
        n,
        "Commitment to Clients",
        "Band",
        "",
        "Eyebrow",
        F_EYEBROW,
        "Our Promise",
    )
    add_row(rows, p, n, "Commitment to Clients", "Band", "", "Title", F_H2_HOME, "Commitment to Clients")
    add_row(
        rows,
        p,
        n,
        "Commitment to Clients",
        "Band",
        "",
        "Text",
        F_BODY,
        "We deliver solutions that go beyond compliance—improving financial efficiency, reducing risks, and supporting strategic decision-making for long-term client success.",
    )
    add_row(rows, p, n, "Our Presence", "Cards", "", "Eyebrow", F_EYEBROW, "Where We Are")
    add_row(rows, p, n, "Our Presence", "Cards", "", "Title", F_H2_PRESENCE, "Our Presence")
    add_row(
        rows,
        p,
        n,
        "Our Presence",
        "Cards",
        "",
        "Intro",
        F_BODY,
        "Headquartered in Varanasi with branch offices across key cities to serve you locally.",
    )
    add_row(
        rows,
        p,
        n,
        "Our Presence",
        "HQ card",
        "",
        "Head Office – Varanasi",
        F_BODY,
        "Our flagship office and central hub for assurance, tax, and advisory services. (Card shows city Varanasi; full address in footer / contact.)",
    )
    for city, addr, contact in [
        ("Delhi", "62, Shrestha Vihar, Vikas Marg Extension, Delhi – 110092", "In charge: Delhi Branch. Reach us for audit, tax & compliance in NCR."),
        (
            "Kolkata",
            "Brijdham Housing Complex, 637 Dakshin Dari Road, 5th Floor Flat-5E, Building No 16-C, Kolkata, West Bengal",
            "In charge: Kolkata Branch. Serving Eastern India with local expertise.",
        ),
        (
            "Bokaro",
            "C-1, 21A, 2nd Floor, City Centre, Sector-4, Bokaro Steel City, Jharkhand",
            "In charge: Bokaro Branch. Industry-focused CA services in the region.",
        ),
    ]:
        add_no = "Branch card"
        add_row(rows, p, n, "Our Presence", add_no, "", f"Branch – {city}", F_BODY, f"{addr}\n{contact}")
    add_row(
        rows,
        p,
        n,
        "Our Presence",
        "Cards",
        "",
        "Tagline",
        F_BODY,
        "We serve clients across India with strong regional expertise and national capabilities.",
    )
    add_row(rows, p, n, "Bottom CTA", "Gradient band", "Schedule a Consultation", "Eyebrow", F_EYEBROW, "Expert guidance since 2003")
    add_row(rows, p, n, "Bottom CTA", "Gradient band", "", "Title", F_CTA_BAND, "Partner with Dwivedi Gupta & Co.")
    add_row(
        rows,
        p,
        n,
        "Bottom CTA",
        "Gradient band",
        "",
        "Description",
        F_BODY,
        "For trusted financial guidance and professional excellence.",
    )

    # ----- SERVICES (3) -----
    p, n = "SERVICES", 3
    add_row(rows, p, n, "Hero", "Hero", "Schedule Consultation | Contact Us", "Label", F_LABEL, "DWIVEDI GUPTA & CO.")
    add_row(rows, p, n, "Hero", "Hero", "", "Tagline", F_TAG, "Chartered Accountants | Assurance | Tax | Advisory")
    add_row(rows, p, n, "Hero", "Hero", "", "Title", F_H1_HERO, "Our Services")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Hero",
        "",
        "Description",
        F_HERO_DESC,
        "End-to-end chartered accountancy services for businesses and individuals—taxation, audit & assurance, corporate law, financial consulting, and government schemes advisory.",
    )
    add_row(
        rows,
        p,
        n,
        "Overview",
        "Text block",
        "",
        "Paragraph",
        F_BODY,
        "Dwivedi Gupta & Co. has been delivering assurance, taxation, advisory, and consulting services since 2003. Our partner-led, client-centric approach ensures that every engagement is tailored to your business needs. Whether you are a growing SME, a listed company, or a public-sector entity, we combine technical expertise with practical insights to help you stay compliant, manage risk, and achieve your financial goals.",
    )
    add_row(rows, p, n, "Why services", "Grid", "", "Eyebrow", F_EYEBROW, "Why Choose Us")
    add_row(rows, p, n, "Why services", "Grid", "", "Title", F_H2_HOME, "Why Choose Our Services")
    add_row(
        rows,
        p,
        n,
        "Why services",
        "Grid",
        "",
        "Intro",
        F_BODY,
        "We combine technical excellence with a commitment to long-term partnerships. Here is what sets our service delivery apart.",
    )
    for title, desc in [
        (
            "Partner-Led Engagement",
            "Senior partners are directly involved in your engagement, ensuring quality and accountability.",
        ),
        (
            "Full-Service Under One Roof",
            "From incorporation and tax to audit and advisory—we cover the full lifecycle so you have one trusted partner.",
        ),
        (
            "Timely & Reliable Delivery",
            "We meet deadlines and keep you informed so you never miss a compliance date or opportunity.",
        ),
        (
            "RBI & CAG Empanelled",
            "Our empanelment reflects the standards we uphold and our capability to serve institutions.",
        ),
        (
            "Multi-City Presence",
            "Head office in Varanasi with branches in Delhi, Kolkata, and Bokaro for nationwide reach.",
        ),
        (
            "Industry & Sector Expertise",
            "We serve manufacturing, banking, real estate, government, SMEs, and non-profits with tailored approaches.",
        ),
    ]:
        add_row(rows, p, n, "Why services", "Card", "", title, F_BODY, desc)
    for step, title, text in [
        ("01", "Consult", "We understand your business, goals, and challenges through a structured consultation."),
        ("02", "Plan", "We design a clear scope, timeline, and deliverables aligned with your requirements."),
        ("03", "Deliver", "Our team executes with quality and keeps you updated at every stage."),
        ("04", "Support", "Ongoing support and reviews so you stay compliant and informed."),
    ]:
        add_row(rows, p, n, "Our Approach", "Step", "", f"{step} {title}", F_BODY, text)
    add_row(rows, p, n, "Service grid", "Linked cards", "Schedule | Discuss", "Section", F_H2_HOME, "Comprehensive CA & Advisory Solutions")
    for svc in build_services_data():
        add_row(rows, p, n, "Service grid", "Card", "Learn more →", "Title", F_H3_CARD, svc["title"])
        add_row(rows, p, n, "Service grid", "Card", "", "Short description", F_BODY, svc["short"])
        add_row(rows, p, n, "Service grid", "Card", "", "Key areas (first 3 on card)", F_BODY, "; ".join(svc["areas"].split("; ")[:3]))

    add_row(rows, p, n, "Detailed catalogue", "Articles", "", "Eyebrow", F_EYEBROW, "Detailed Services")
    add_row(rows, p, n, "Detailed catalogue", "Articles", "", "Title", F_H2_HOME, "Detailed Service Catalogue")
    add_row(
        rows,
        p,
        n,
        "Detailed catalogue",
        "Articles",
        "",
        "Intro",
        F_BODY,
        "Explore each service in depth, including scope, key focus areas, and the type of clients it is best suited for.",
    )
    for svc in build_services_data():
        slug = svc["slug"]
        aud = SERVICE_AUDIENCE.get(
            slug,
            "Suitable for businesses and professionals seeking structured compliance and advisory support.",
        )
        add_row(rows, p, n, "Detailed catalogue", "Article", "Open detailed page →", "Service title", F_H3_CARD, svc["title"])
        add_row(rows, p, n, "Detailed catalogue", "Article", "", "Summary", F_BODY, svc["short"])
        add_row(rows, p, n, "Detailed catalogue", "Article", "", "Body", F_BODY, svc["long"])
        add_row(rows, p, n, "Detailed catalogue", "Article", "", "Key deliverables", F_BODY, svc["areas"])
        add_row(rows, p, n, "Detailed catalogue", "Article", "", "Who this is for", F_BODY, aud)

    # ----- INDUSTRIES (4) -----
    p, n = "INDUSTRIES", 4
    add_row(rows, p, n, "Hero", "Hero", "Our Services | Contact Us", "Label", F_LABEL, "DWIVEDI GUPTA & CO.")
    add_row(rows, p, n, "Hero", "Hero", "", "Tagline", F_TAG, "Sector Expertise · Compliance · Advisory")
    add_row(rows, p, n, "Hero", "Hero", "", "Title", F_H1_HERO, "Industries We Serve")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Hero",
        "",
        "Description",
        F_HERO_DESC,
        "We provide assurance, tax, and advisory services across multiple industries with sector-specific knowledge and compliance support tailored to your business.",
    )
    add_row(rows, p, n, "Intro", "Section", "", "Eyebrow", F_EYEBROW, "Sector Expertise")
    add_row(rows, p, n, "Intro", "Section", "", "H2", F_H2_HOME, "Industry-Focused Solutions")
    add_row(
        rows,
        p,
        n,
        "Intro",
        "Section",
        "",
        "Lead",
        F_BODY,
        "Every industry has unique compliance needs, reporting standards, and growth challenges. Dwivedi Gupta & Co. combines deep technical expertise with sector knowledge to deliver practical, timely, and value-driven solutions.",
    )
    add_row(
        rows,
        p,
        n,
        "Intro",
        "Section",
        "",
        "Paragraph",
        F_BODY,
        "From manufacturing and banking to real estate, trading, government, SMEs, and non-profits—we serve a diverse client base across India. Our partner-led approach ensures that you receive advice that understands your sector’s regulations, risks, and opportunities.",
    )
    industry_cards = [
        (
            "Manufacturing & Infrastructure",
            "We support manufacturers and infrastructure players with cost accounting, tax incentives, GST compliance, and project finance.",
            "Tax incentives & exemptions; Cost accounting & MIS; GST and indirect tax compliance; Statutory & internal audit",
        ),
        (
            "Banking & Financial Institutions",
            "RBI compliance, statutory audit, internal audit, and risk assurance for banks, NBFCs, and financial institutions.",
            "RBI compliance & reporting; Statutory & branch audits; Internal audit & risk; Tax and transfer pricing",
        ),
        (
            "Real Estate & Construction",
            "Project accounting, GST on construction, RERA compliance, and regulatory support for developers and contractors.",
            "Project & contract accounting; GST on works contract; RERA and regulatory filings; Joint venture structuring",
        ),
        (
            "Trading & Export Businesses",
            "Customs, export benefits, multi-state GST, and inventory accounting for trading and export-oriented businesses.",
            "Export incentives & MEIS/SEIS; Customs and DGFT compliance; Multi-state GST registration; Inventory & working capital",
        ),
        (
            "Government & Public Sector",
            "CAG empanelment, government audits, and compliance support for public sector undertakings and government schemes.",
            "CAG & government audits; Scheme implementation support; Compliance & reporting; Procurement and contracts",
        ),
        (
            "SMEs & Startups",
            "Company formation, ESOP accounting, startup tax benefits, and compliance tailored for growing and early-stage businesses.",
            "Incorporation & secretarial; Income tax 80-IAC & incentives; ESOP & ESOP valuation; GST and annual compliance",
        ),
        (
            "Non-Profit Organizations",
            "Trust and society compliance, 80G/12A support, fund accounting, and audit for NGOs and non-profit entities.",
            "Trust/society compliance; 80G, 12A & FCRA; Fund accounting & audit; Board and donor reporting",
        ),
    ]
    for title, desc, svcs in industry_cards:
        add_row(rows, p, n, "Sectors", "Sector card", "", "Sector title", F_H3_CARD, title)
        add_row(rows, p, n, "Sectors", "Sector card", "", "Description", F_BODY, desc)
        add_row(rows, p, n, "Sectors", "Sector card", "", "Services list", F_BODY, svcs)
    add_row(rows, p, n, "Why sector", "Bullets", "", "Point 1", F_BODY, "Regulatory requirements vary significantly by sector—we bring sector-specific knowledge to ensure full compliance.")
    add_row(rows, p, n, "Why sector", "Bullets", "", "Point 2", F_BODY, "Industry benchmarks and best practices help us deliver more relevant advice and efficient audits.")
    add_row(rows, p, n, "Why sector", "Bullets", "", "Point 3", F_BODY, "We understand your value chain and key risks, so our solutions are practical and aligned with your operations.")
    add_row(rows, p, n, "CTA", "Band", "Schedule a Consultation | Contact Us", "Eyebrow", F_EYEBROW, "Expert guidance since 2003")
    add_row(
        rows,
        p,
        n,
        "CTA",
        "Band",
        "",
        "Title",
        F_CTA_BAND,
        "Your Industry. Our Expertise.",
    )
    add_row(
        rows,
        p,
        n,
        "CTA",
        "Band",
        "",
        "Description",
        F_BODY,
        "Partner with Dwivedi Gupta & Co. for sector-specific assurance, tax, and advisory support.",
    )

    # ----- TEAM (5) -----
    p, n = "TEAM", 5
    add_row(rows, p, n, "Hero", "Page hero", "Schedule Consultation | Contact Us", "Title", F_PAGE_H1, "Our Team")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Page hero",
        "",
        "Subtitle",
        F_PAGE_SUB,
        "Under the able direction of 7 partners and other qualified CA staff, DGC's team is uniquely positioned to provide you quality opinions and services.",
    )
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Page hero",
        "",
        "Note",
        F_BODY,
        "Click View more on any profile to see full qualifications, experience, and contact details.",
    )
    partners = [
        "CA. Surendra Kumar Dwivedi — Founder Partner — FCA, FAFP, DISA, B.Com, LLB — Direct & Indirect Tax, Assurance & Accounting, Project Finance/Debt Syndication — 33 Years",
        "CA. Vivek Anand Mohan — Partner — FCA, B.S.C., LLB, DISA(ICAI), DIM, CCCAB — Assurance & Accounting — 18 Years",
        "CA. Aditi Kapoor — Partner — FCA, DISA (ICAI), DIRM (ICAI), B.Com (H) — Financial Advisory, Technical & Economic Viability, Treasury, Direct Tax, Government Schemes — 14 Years — In Charge: Delhi Branch",
        "CA. Sharad Kumar Jaiswal — Partner — FCA, B.Com — Audit, Project financing, Statutory Compliance — 16 Years",
        "CA. Shweta Bharuka — Partner — ACA, B. Com (H) — Direct Tax, Assurance & Accounting — 9 Years — In-charge: Kolkata Branch",
        "CA. Neha Nathani — Partner — FCA and B.Com (H) — Taxation and Auditing — 11 Years",
        "CA. Breesket Singh — Partner — FCA and B.Com (H) — Taxation and Auditing — 21 Years — In Charge: Bokaro Branch",
    ]
    for line in partners:
        add_row(rows, p, n, "Team grid", "Card", "View more", "Member preview", F_BODY, line)
    add_row(
        rows,
        p,
        n,
        "CTA",
        "Buttons",
        "Careers | Get in Touch",
        "Text",
        F_BODY,
        "We recruit, train, motivate and retain highly capable talent who deliver the best solutions. Want to work with us?",
    )

    # ----- CLIENTS (6) -----
    p, n = "CLIENTS", 6
    add_row(rows, p, n, "Hero", "Page hero", "", "Title", F_PAGE_H1, "Our Clients")
    add_row(rows, p, n, "Hero", "Page hero", "", "Subtitle", F_PAGE_SUB, "We serve corporates, MSMEs, startups, and individuals across sectors.")
    add_row(rows, p, n, "Logos", "Marquee", "", "Section title", F_SECTION_H2, "Our Clients")
    add_row(
        rows,
        p,
        n,
        "Logos",
        "Marquee",
        "",
        "Subtitle",
        F_BODY,
        "We are proud to work with businesses and individuals across sectors.",
    )
    add_row(rows, p, n, "Logos", "Marquee", "", "Alt text", F_BODY, "Logo strip: Union Bank of India, SBI, Punjab National Bank, Bank of Baroda, Kajaria Tiles, Haldiram's, Intex, Akshat Group, RC Rungta Group, Eco Plus, RLJ Group (see clientLogos.js).")
    add_row(rows, p, n, "Who we work with", "Cards", "", "Title", F_SECTION_H2, "Who We Work With")
    add_row(
        rows,
        p,
        n,
        "Who we work with",
        "Cards",
        "",
        "Subtitle",
        F_BODY,
        "From large corporates to individuals—we tailor our services to your size and needs.",
    )
    for title, desc in [
        ("Corporates & Listed Companies", "Statutory audit, tax planning, and compliance for listed and unlisted corporates."),
        ("MSMEs & Growing Businesses", "Bookkeeping, GST, and annual compliance tailored for small and medium enterprises."),
        ("Startups", "Incorporation, funding compliance, and startup-specific tax and regulatory support."),
        ("Individuals & HNIs", "Personal tax planning, ITR filing, and wealth management compliance."),
    ]:
        add_row(rows, p, n, "Who we work with", "Card", "", title, F_BODY, desc)
    add_row(rows, p, n, "Testimonials", "Quotes", "", "Title", F_SECTION_H2, "What Clients Say")
    for q, a in [
        ("Professional, timely, and always available. Our go-to firm—DWIVEDI GUPTA & Co.—for the last five years.", "Client, Manufacturing"),
        ("They simplified our GST and compliance. Highly recommend for startups.", "Client, Technology"),
    ]:
        add_row(rows, p, n, "Testimonials", "Quotes", "", "Quote", F_BODY, f'"{q}" — {a}')
    add_row(rows, p, n, "Relationships", "Stats", "", "Title", F_SECTION_H2, "Long-Term Client Relationships")
    add_row(
        rows,
        p,
        n,
        "Relationships",
        "Stats",
        "",
        "Subtitle",
        F_BODY,
        "Built through consistent quality, transparent communication, and practical advisory.",
    )
    for val, lab in [
        ("92%+", "Client Retention"),
        ("15+", "Industries Served"),
        ("4 Offices", "Multi-City Support"),
        ("< 24 Hrs", "Average Response Time"),
    ]:
        add_row(rows, p, n, "Relationships", "Stat", "", lab, F_BODY, val)
    add_row(rows, p, n, "How we engage", "Steps", "", "Title", F_SECTION_H2, "How We Engage")
    add_row(
        rows,
        p,
        n,
        "How we engage",
        "Steps",
        "",
        "Subtitle",
        F_BODY,
        "A clear process that keeps your business informed, compliant, and audit-ready.",
    )
    for title, desc in [
        ("Discovery & Requirement Mapping", "We understand your business model, compliance pain points, timelines, and reporting expectations."),
        ("Custom Engagement Plan", "A practical scope and service plan is prepared with clear deliverables, ownership, and milestones."),
        ("Execution & Compliance Monitoring", "Our team executes with regular checkpoints, timely filings, and proactive updates for leadership."),
        ("Review & Ongoing Advisory", "We close each cycle with performance review, improvement pointers, and year-round advisory support."),
    ]:
        add_row(rows, p, n, "How we engage", "Step", "", title, F_BODY, desc)
    add_row(rows, p, n, "FAQ", "Accordion", "", "Title", F_SECTION_H2, "Frequently Asked Questions")
    for q, a in [
        ("Do you work with businesses outside your office cities?", "Yes. We serve clients across India using a hybrid model of remote collaboration and scheduled on-site support."),
        ("Can you handle both audit and tax for the same organization?", "Yes. Depending on engagement requirements and independence norms, we support both functions through dedicated teams."),
        ("What is the usual onboarding timeline?", "Most engagements are onboarded within 3-7 working days after scope finalization and documentation."),
    ]:
        add_row(rows, p, n, "FAQ", "Item", "", f"Q: {q}", F_BODY, a)
    add_row(rows, p, n, "CTA", "Section", "Start a Conversation", "Message", F_BODY, "Join our growing list of satisfied clients.")

    # ----- INSIGHTS list (7) -----
    p, n = "INSIGHTS", 7
    add_row(rows, p, n, "Hero", "Hero", "Schedule Consultation | Contact Us", "Label", F_LABEL, "DWIVEDI GUPTA & CO.")
    add_row(rows, p, n, "Hero", "Hero", "", "Tagline", F_TAG, "Updates on tax, compliance, and regulatory matters")
    add_row(rows, p, n, "Hero", "Hero", "", "Title", F_H1_HERO, "Insights & Blog")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Hero",
        "",
        "Description",
        F_HERO_DESC,
        "Practical articles and updates on GST, tax, company law, compliance, and advisory—to help you stay informed and compliant.",
    )
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Hero",
        "Schedule Consultation | Contact Us",
        "CTAs",
        F_BTN,
        "Primary: Schedule Consultation — Outline: Contact Us",
    )
    add_row(
        rows,
        p,
        n,
        "Intro",
        "Paragraph",
        "",
        "Body",
        F_BODY,
        "Our insights cover GST, direct and indirect tax, MCA/ROC compliance, audit, startup and MSME schemes, and general advisory. We publish updates and checklists to keep you ahead of due dates and regulatory changes. For advice tailored to your situation, get in touch.",
    )
    add_row(rows, p, n, "Topics", "Filter buttons", "", "Eyebrow", F_EYEBROW, "Browse by topic")
    add_row(rows, p, n, "Topics", "Filter buttons", "", "Title", F_H2_HOME, "Topics We Cover")
    add_row(rows, p, n, "Topics", "Filter buttons", "", "Categories", F_BODY, "All | Tax | GST | Compliance | Audit | Startups | Advisory | Government Schemes")
    for art in build_insights():
        add_row(
            rows,
            p,
            n,
            "Article list",
            "Card",
            "",
            "Article",
            F_H3_CARD,
            f"{art['title']} — {art['cat']} — {art['date']}\nSummary: {art['summary']}\n\nBody: {art['body']}",
        )
    add_row(
        rows,
        p,
        n,
        "Bottom CTA",
        "Banner",
        "Contact Us | Schedule Consultation",
        "Text",
        F_BODY,
        "Need advice on a specific topic? We can help.",
    )
    add_row(
        rows,
        p,
        n,
        "Insights UI",
        "Empty state",
        "",
        "Note",
        F_BODY,
        'When a category has no posts: "No articles in this category yet. Check back soon or browse \\"All\\"."',
    )

    # ----- CAREERS (8) -----
    p, n = "CAREERS", 8
    add_row(rows, p, n, "Main", "Page hero + text", "Contact Us", "Title", F_PAGE_H1, "Careers")
    add_row(
        rows,
        p,
        n,
        "Main",
        "Page hero + text",
        "",
        "Body",
        F_BODY,
        "We are always looking for talented chartered accountants and finance professionals. Email vivek.gupta@dgc.ind.in",
    )

    # ----- SCHEDULE CONSULTATION (9) -----
    p, n = "SCHEDULE CONSULTATION", 9
    add_row(rows, p, n, "Hero", "sched-hero", "", "Eyebrow", F_SCHED, "Schedule a Consultation")
    add_row(rows, p, n, "Hero", "sched-hero", "", "Title", F_SCHED, "Book a Free Consultation")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "sched-hero",
        "",
        "Description",
        F_SCHED,
        "Reserve a 30-minute slot with our CAs. Choose your preferred date and time—we’ll confirm within 1–2 business days. No obligation.",
    )
    add_row(
        rows,
        p,
        n,
        "Form",
        "Form",
        "Submit request",
        "Fields",
        F_SCHED,
        "Full name*, Email*, Phone*, Service of interest (dropdown: Tax Planning & Compliance | Audit & Assurance | Accounting & Bookkeeping | GST Advisory | Company Formation & Compliance | Financial Advisory | Other), Preferred date, Preferred time, Message. Time slots: 9–10 AM through 4–5 PM as in codebase.",
    )
    add_row(
        rows,
        p,
        n,
        "Form",
        "Success",
        "",
        "Message",
        F_BODY,
        "Thank you. We will get back to you soon. / Your consultation request has been received. We will confirm your slot shortly.",
    )

    # ----- CONTACT (10) -----
    p, n = "CONTACT", 10
    add_row(rows, p, n, "Hero", "contact-hero", "", "Eyebrow", F_EYEBROW, "Get in Touch")
    add_row(rows, p, n, "Hero", "contact-hero", "", "Title", F_PAGE_H1, "Contact Us")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "contact-hero",
        "",
        "Intro",
        F_BODY,
        "Whether you need tax advice, audit support, GST compliance, or company formation—our team is here to help. Reach out via the form, email, or phone. We respond to all enquiries within 24–48 hours.",
    )
    add_row(rows, p, n, "Sidebar", "Card", "", "Title", F_SECTION_H2, "Reach Us Directly")
    add_row(
        rows,
        p,
        n,
        "Sidebar",
        "Card",
        "",
        "Description",
        F_BODY,
        "For urgent matters or a quick conversation, use the details below.",
    )
    add_row(
        rows,
        p,
        n,
        "Sidebar",
        "Card",
        "",
        "Email",
        F_BODY,
        "vivek.gupta@dgc.ind.in; shikhar.dwivedi@dgc.ind.in",
    )
    add_row(
        rows,
        p,
        n,
        "Sidebar",
        "Card",
        "",
        "Phone",
        F_BODY,
        "+91 94158 05906 | +91 9721227799",
    )
    add_row(rows, p, n, "Sidebar", "Card", "", "WhatsApp", F_BODY, "Chat with us (wa.me primary number)")
    add_row(rows, p, n, "Sidebar", "Card", "", "Offices heading", F_SECTION_H2, "Our Offices")
    add_row(
        rows,
        p,
        n,
        "Sidebar",
        "Card",
        "",
        "Offices intro",
        F_BODY,
        "Head office in Varanasi and branch offices across India. Visit by appointment.",
    )
    for block in [
        "Varanasi — Head Office — S-8/108-B-3-A Prashantpuri, M.A Road — flagship; visit by appointment — Get directions →",
        "Delhi — Branch — 62, Shrestha Vihar, Vikas Marg Extension — NCR — Get directions →",
        "Kolkata — Branch — Brijdham Housing Complex, 637 Dakshin Dari Road, Flat-5E, Bldg 16-C — Get directions →",
        "Bokaro — Branch — C-1, 21A, 2nd Floor, City Centre, Sector-4 — Get directions →",
    ]:
        add_row(rows, p, n, "Sidebar", "Office item", "", "Office", F_BODY, block)
    add_row(rows, p, n, "Sidebar", "Expect", "", "Title", F_SECTION_H2, "What to Expect")
    for line in [
        "Response within 24–48 hours on working days",
        "Confidentiality on all enquiries",
        "Support in Tax, Audit, GST, Company Formation, and more",
        "Option to schedule a consultation for detailed discussions",
    ]:
        add_row(rows, p, n, "Sidebar", "Expect", "Schedule a Consultation", "Bullet", F_BODY, line)
    add_row(rows, p, n, "Form", "Form", "Send Message", "Title", F_SECTION_H2, "Send a Message")
    add_row(
        rows,
        p,
        n,
        "Form",
        "Form",
        "",
        "Description",
        F_BODY,
        "Fill in the form below and we’ll get back to you as soon as possible.",
    )
    add_row(
        rows,
        p,
        n,
        "Form",
        "Form",
        "Send Message",
        "Fields",
        F_BODY,
        "Name*, Email*, Phone, Subject (General Enquiry, Tax, Audit, GST, Company Formation, Compliance, Other), Message*",
    )
    add_row(
        rows,
        p,
        n,
        "Form",
        "Form",
        "",
        "Success",
        F_BODY,
        "Thank you. We will get back to you within 24–48 hours.",
    )

    # ----- COMPLIANCE (11) -----
    p, n = "COMPLIANCE", 11
    add_row(rows, p, n, "Main", "Page", "", "Title", F_PAGE_H1, "Compliance Information")
    add_row(
        rows,
        p,
        n,
        "Main",
        "Page",
        "",
        "Subtitle",
        F_PAGE_SUB,
        "Regulatory and compliance information for Dwivedi Gupta & Co.",
    )
    for label, val in [
        ("Year of Establishment", "2003"),
        ("Firm Registration No.", "012584C"),
        ("RBI Unique Code", "000293"),
        ("Peer Review Certificate No.", "014832"),
        ("CAG Empanelment No.", "CR3209"),
        ("Category of Firm (ICAI)", "I (One)"),
        ("PAN No.", "AAEFD6000D"),
        ("GSTN No.", "09AAEFD6000D1Z1"),
        ("Official Email", "aditi.kapoor@dgc.ind.in"),
        ("Head Office", "Varanasi"),
        ("Branch Offices", "Kolkata, Delhi & Bokaro"),
    ]:
        add_row(rows, p, n, "Credentials", "Grid", "", label, F_COMPLIANCE, val)

    # ----- LEGAL SHORT PAGES -----
    p, n = "PRIVACY POLICY", 12
    add_row(rows, p, n, "Hero", "Page hero", "", "Title", F_PAGE_H1, "Privacy Policy")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Page hero",
        "",
        "Subtitle",
        F_PAGE_SUB,
        "How Dwivedi Gupta & Co. collects, uses, and protects your information.",
    )
    add_row(rows, p, n, "Main", "Page", "", "Body", F_BODY, "This page will contain our privacy policy. Last updated: 2026.")
    p, n = "TERMS", 13
    add_row(rows, p, n, "Hero", "Page hero", "", "Title", F_PAGE_H1, "Terms & Conditions")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Page hero",
        "",
        "Subtitle",
        F_PAGE_SUB,
        "Terms of use for the Dwivedi Gupta & Co. website and services.",
    )
    add_row(rows, p, n, "Main", "Page", "", "Body", F_BODY, "This page will contain our terms and conditions. Last updated: 2026.")
    p, n = "DISCLAIMER", 14
    add_row(rows, p, n, "Hero", "Page hero", "", "Title", F_PAGE_H1, "Disclaimer")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Page hero",
        "",
        "Subtitle",
        F_PAGE_SUB,
        "Legal disclaimer for information provided by Dwivedi Gupta & Co.",
    )
    add_row(
        rows,
        p,
        n,
        "Main",
        "Page",
        "",
        "Paragraph 1",
        F_BODY,
        "The information contained in this document prepared by Dwivedi Gupta & Co. (hereinafter referred to as DGC) is furnished to the recipient, on his/ her specific request and for information purpose only. In no way, this document should be treated as a marketing material or efforts to solicit a client. The sole purpose of this document is to furnish factual information about DGC's profile.",
    )
    add_row(
        rows,
        p,
        n,
        "Main",
        "Page",
        "",
        "Paragraph 2",
        F_BODY,
        "While we have made every attempt to ensure that the information contained in this document is true, DGC, its partners and/or any of its employees does not give any warranty, express or implied, including the warranty of opinions expressed for a particular purpose, or assume any liability or responsibility for the accuracy, completeness, or usefulness of any information available from this document.",
    )
    add_row(rows, p, n, "Main", "Page", "", "Updated", F_BODY, "Last updated: 2025.")

    p, n = "SITEMAP", 15
    add_row(rows, p, n, "Hero", "Page hero", "", "Title", F_PAGE_H1, "Sitemap")
    add_row(
        rows,
        p,
        n,
        "Hero",
        "Page hero",
        "",
        "Subtitle",
        F_PAGE_SUB,
        "All pages on the Dwivedi Gupta & Co. website.",
    )
    add_row(
        rows,
        p,
        n,
        "Main",
        "Links",
        "",
        "Description",
        F_BODY,
        "Index of all primary routes, service pages, and insight articles (see Sitemap.jsx).",
    )

    # ----- SERVICE DETAIL PAGES (16) -----
    p, n = "SERVICE DETAIL (×8)", 16
    for svc in build_services_data():
        add_row(
            rows,
            p,
            n,
            svc["slug"],
            "Hero + body",
            "Schedule Consultation | Contact Us",
            "H1",
            F_H1_HERO_SM,
            svc["title"],
        )
        add_row(rows, p, n, svc["slug"], "Hero + body", "", "Short", F_HERO_DESC, svc["short"])
        add_row(rows, p, n, svc["slug"], "Overview", "", "Long", F_BODY, svc["long"])
        add_row(rows, p, n, svc["slug"], "What We Offer", "", "Key areas", F_BODY, svc["areas"])

    # ----- FOOTER -----
    p, n = "FOOTER (all pages)", 17
    add_row(rows, p, n, "Brand column", "Footer", "Schedule Consultation", "Company name", F_H3_CARD, "Dwivedi Gupta & Co.")
    add_row(
        rows,
        p,
        n,
        "Brand column",
        "Footer",
        "",
        "Tagline",
        F_BODY,
        "Dwivedi Gupta & Co. is a Chartered Accountants firm providing taxation, audit, advisory, and financial consulting services since 2003 with offices in Varanasi, Delhi, Kolkata, and Bokaro.",
    )
    add_row(
        rows,
        p,
        n,
        "Brand column",
        "Footer",
        "",
        "Social",
        F_BODY,
        "LinkedIn (company page) | mailto:vivek.gupta@dgc.ind.in | tel:+919415805906 | WhatsApp float",
    )
    add_row(
        rows,
        p,
        n,
        "Quick links",
        "Footer",
        "",
        "Links",
        F_BODY,
        "Home | About Us | Services | Industries | Team | Clients | Insights | Careers | Contact | Schedule Consultation",
    )
    add_row(
        rows,
        p,
        n,
        "Services links",
        "Footer",
        "",
        "All service pages",
        F_BODY,
        "Tax & Compliance | Audit & Assurance | GST Advisory | Corporate Law | Company Formation | Financial Consulting | Project Finance | Government Schemes Advisory",
    )
    add_row(
        rows,
        p,
        n,
        "Offices",
        "Footer",
        "",
        "Addresses",
        F_BODY,
        "Varanasi (Head); Delhi; Kolkata; Bokaro — full addresses as in Footer.jsx",
    )
    add_row(rows, p, n, "Newsletter", "Footer", "Subscribe", "Heading", F_BODY, "Newsletter")
    add_row(rows, p, n, "Newsletter", "Footer", "Subscribe", "Placeholder", F_BODY, "Your email")
    add_row(rows, p, n, "Newsletter", "Footer", "", "Success", F_BODY, "Thank you for subscribing.")
    add_row(
        rows,
        p,
        n,
        "Resources",
        "Footer",
        "",
        "Links",
        F_BODY,
        "Sitemap | Privacy Policy | Terms & Conditions | Disclaimer | Compliance Information",
    )
    add_row(
        rows,
        p,
        n,
        "Bottom bar",
        "Footer",
        "",
        "Copyright",
        F_BODY,
        "© [current year] Dwivedi Gupta & Co. All Rights Reserved. — Privacy | Terms | Sitemap",
    )
    add_row(
        rows,
        p,
        n,
        "Floating",
        "UI",
        "",
        "WhatsApp + Scroll top",
        F_BODY,
        "Fixed WhatsApp button; scroll-to-top control",
    )

    return rows


def write_workbook(rows: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Audit"

    ws.merge_cells("G1:I1")
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:D2")
    ws.merge_cells("E1:E2")
    ws.merge_cells("F1:F2")
    ws.merge_cells("J1:J2")

    headers_top = [
        ("A1", "S.no"),
        ("B1", "Page Name"),
        ("C1", "Page No"),
        ("D1", "Page Sub-heading"),
        ("E1", "Design Used"),
        ("F1", "Buttons"),
        ("G1", "Content"),
        ("J1", "Remarks"),
    ]
    for coord, val in headers_top:
        ws[coord] = val

    ws["G2"] = "Heading"
    ws["H2"] = "Font name & Size"
    ws["I2"] = "Text"

    header_font = Font(name="Times New Roman", bold=True, size=11)
    title_fill = PatternFill("solid", fgColor="D9E1F2")
    for r in (1, 2):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            if cell.value:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = title_fill
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    data_font = Font(name="Times New Roman", size=10)
    meta_font = Font(name="Times New Roman", size=9, italic=True)

    start_row = 3
    for i, r in enumerate(rows, start=0):
        row_idx = start_row + i
        ws.cell(row=row_idx, column=1, value=i + 1)
        ws.cell(row=row_idx, column=2, value=r["page_name"])
        ws.cell(row=row_idx, column=3, value=r["page_no"])
        ws.cell(row=row_idx, column=4, value=r["sub"])
        ws.cell(row=row_idx, column=5, value=r["design"])
        ws.cell(row=row_idx, column=6, value=r["buttons"])
        ws.cell(row=row_idx, column=7, value=r["heading"])
        ws.cell(row=row_idx, column=8, value=r["font"])
        ws.cell(row=row_idx, column=9, value=r["text"])
        ws.cell(row=row_idx, column=10, value=r["remarks"])

        for c in range(1, 11):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = data_font if c != 8 else meta_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    if rows:
        ws.cell(
            row=start_row,
            column=10,
            value="Column H = typography on the live site (DM Sans). Table body uses Times New Roman to match typical audit sheets.",
        )

    # Vertical merges for Page Name (B) and Page No (C)
    if rows:
        def merge_page_groups(col_idx: int):
            i = 0
            col_letter = get_column_letter(col_idx)
            while i < len(rows):
                pname = rows[i]["page_name"]
                pno = rows[i]["page_no"]
                j = i
                while j < len(rows) and rows[j]["page_name"] == pname and rows[j]["page_no"] == pno:
                    j += 1
                if j - i > 1:
                    r1 = start_row + i
                    r2 = start_row + j - 1
                    ws.merge_cells(f"{col_letter}{r1}:{col_letter}{r2}")
                i = j

        merge_page_groups(2)
        merge_page_groups(3)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 36
    ws.column_dimensions["I"].width = 72
    ws.column_dimensions["J"].width = 16

    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    wb.save(OUT)
    print(f"Wrote {OUT} ({len(rows)} data rows)")


def main():
    data = build_all_rows()
    write_workbook(data)


if __name__ == "__main__":
    main()
